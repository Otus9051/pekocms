import sys, os, json, webbrowser, datetime
from typing import List, Dict, Any
from PySide6 import QtWidgets, QtCore, QtGui, QtSvg

# Add parent directory to path for db imports
# Add parent directory to path for db imports
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)

from db import auth_db, patient_cms_db, datasheet_db, report_tracker_db, invoice_service, data_fetcher, catalogue_db, special_tests_db, polyclinic_db
from app import theme
# Import refactored modules
from app.login import LoginWindow
from app.threads import (
    CatalogueLoaderThread, 
    InvoiceCatalogueLoaderThread, 
    SpecialTestsLoaderThread, 
    FullCatalogueLoaderThread
)
from app.updater import check_for_updates_gui
from app.branding import (
    APP_NAME, LOGIN_WINDOW_TITLE, LOGIN_WINDOW_HEADING,
    CLINIC_NAME, CLINIC_ADDRESS, CLINIC_CONTACT,
    WINDOW_TITLE_PATHOLOGY, WINDOW_TITLE_POLYCLINIC, WINDOW_TITLE_ADMIN,
    LOGO_SVG, LOGO_PNG, ASSETS_DIRECTORY, PATIENT_ID_COLUMN_NAME,
    QUEUE_TABLE_HEADERS, REPORT_TABLE_HEADERS, PATIENT_TABLE_HEADERS,
    MODE_PATHOLOGY, MODE_POLYCLINIC, MODE_ADMIN,
    BUTTON_SIGN_IN, BUTTON_SHUTDOWN, BUTTON_REFRESH, BUTTON_EXPORT,
    DEFAULT_CSV_FILENAME, DEFAULT_XLSX_FILENAME,
    INVOICE_FOLDER_NAME, INVOICE_SUBDIRECTORY, PATIENT_ID_LABEL, CLINIC_NAME_PREFIX,
    FOOTER_TEXT, UI_SCALE
)

from app.utils import get_asset_path, get_invoice_storage_dir, get_config_path
import yaml

# Handle PyInstaller frozen app paths
if getattr(sys, 'frozen', False):
    BASE_PATH = sys._MEIPASS
else:
    BASE_PATH = PROJECT_ROOT

INVOICE_STORAGE_DIR = get_invoice_storage_dir()


def render_svg(svg_path, height):
    """Render an SVG to a QPixmap at specific height for sharpness"""
    if not os.path.exists(svg_path):
        return None
        
    renderer = QtSvg.QSvgRenderer(svg_path)
    if not renderer.isValid():
        return None
        
    size = renderer.defaultSize()
    if size.isEmpty():
        return None
        
    ratio = size.width() / size.height()
    width = int(height * ratio)
    
    # Create high-res pixmap (assuming standard 96dpi, or just render "large enough" for display)
    pixmap = QtGui.QPixmap(width, height)
    pixmap.fill(QtCore.Qt.transparent)
    
    painter = QtGui.QPainter(pixmap)
    painter.setRenderHint(QtGui.QPainter.Antialiasing)
    painter.setRenderHint(QtGui.QPainter.SmoothPixmapTransform)
    renderer.render(painter)
    painter.end()
    
    return pixmap




INVOICE_STORAGE_DIR = get_invoice_storage_dir()


class FirstTimeSetupDialog(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Welcome to {APP_NAME} - Initial Setup")
        self.resize(500, 550)
        self.setup_ui()
        
    def setup_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(40, 40, 40, 40)
        layout.setSpacing(20)
        
        # Branding Header
        header_layout = QtWidgets.QVBoxLayout()
        header_layout.setAlignment(QtCore.Qt.AlignCenter)
        
        # Try to load logo
        logo_label = QtWidgets.QLabel()
        logo_path = get_asset_path(LOGO_SVG)
        pix = render_svg(logo_path, 100) # Crisp render at 100px
        if pix:
            logo_label.setPixmap(pix)
        logo_label.setAlignment(QtCore.Qt.AlignCenter)
        header_layout.addWidget(logo_label)
        
        title = QtWidgets.QLabel(APP_NAME)
        title.setStyleSheet(f"font-size: 28px; font-weight: bold; color: {theme.PRIMARY_COLOR}; margin-top: 15px;")
        title.setAlignment(QtCore.Qt.AlignCenter)
        header_layout.addWidget(title)
        
        subtitle = QtWidgets.QLabel("First Time Setup")
        subtitle.setStyleSheet("font-size: 16px; color: #555; margin-bottom: 20px;")
        subtitle.setAlignment(QtCore.Qt.AlignCenter)
        header_layout.addWidget(subtitle)
        
        layout.addLayout(header_layout)
        
        # --- Option 1: Create Admin ---
        group_box = QtWidgets.QGroupBox("Create Admin Account")
        group_box.setStyleSheet("QGroupBox { font-weight: bold; margin-top: 10px; }")
        group_layout = QtWidgets.QFormLayout(group_box)
        group_layout.setSpacing(12)
        group_layout.setContentsMargins(15, 20, 15, 15)
        
        self.uname = QtWidgets.QLineEdit()
        self.uname.setPlaceholderText("Username")
        self.uname.setStyleSheet("padding: 8px;")
        
        self.pwd = QtWidgets.QLineEdit()
        self.pwd.setPlaceholderText("Password")
        self.pwd.setEchoMode(QtWidgets.QLineEdit.Password)
        self.pwd.setStyleSheet("padding: 8px;")
        
        self.fname = QtWidgets.QLineEdit()
        self.fname.setPlaceholderText("Full Name")
        self.fname.setStyleSheet("padding: 8px;")
        
        group_layout.addRow("Username:", self.uname)
        group_layout.addRow("Password:", self.pwd)
        group_layout.addRow("Full Name:", self.fname)
        
        self.create_btn = QtWidgets.QPushButton("Create Account & Start")
        # Inline style because MainWindow static method might not be available relative to scope or just simpler
        self.create_btn.setStyleSheet(f"background-color: {theme.PRIMARY_COLOR}; color: white; font-weight: bold; padding: 12px; border-radius: 4px; font-size: 14px;")
        self.create_btn.setCursor(QtCore.Qt.PointingHandCursor)
        self.create_btn.clicked.connect(self.create_account)
        group_layout.addRow(self.create_btn)
        
        layout.addWidget(group_box)
        
        # --- Separator ---
        sep = QtWidgets.QLabel("- OR -")
        sep.setAlignment(QtCore.Qt.AlignCenter)
        sep.setStyleSheet("color: #888; font-weight: bold; margin: 10px 0;")
        layout.addWidget(sep)
        
        # --- Option 2: Restore from Backup ---
        restore_frame = QtWidgets.QFrame()
        restore_frame.setStyleSheet("background-color: #f5f5f5; border-radius: 6px; padding: 10px;")
        restore_layout = QtWidgets.QHBoxLayout(restore_frame)
        
        restore_label = QtWidgets.QLabel("Have a backup?")
        restore_label.setStyleSheet("font-weight: 500;")
        
        self.restore_btn = QtWidgets.QPushButton("Restore System from Backup")
        self.restore_btn.setCursor(QtCore.Qt.PointingHandCursor)
        self.restore_btn.setStyleSheet("background-color: #e0e0e0; border: 1px solid #ccc; padding: 6px 12px; border-radius: 4px; font-weight: bold;")
        self.restore_btn.clicked.connect(self.restore_backup)
        
        restore_layout.addWidget(restore_label)
        restore_layout.addStretch()
        restore_layout.addWidget(self.restore_btn)
        
        layout.addWidget(restore_frame)
        layout.addStretch()

    def create_account(self):
        uname = self.uname.text().strip()
        pwd = self.pwd.text().strip()
        fname = self.fname.text().strip()
        
        if not uname or not pwd:
            QtWidgets.QMessageBox.warning(self, "Validation Error", "Username and Password are required.")
            return
            
        try:
            auth_db.create_user(uname, pwd, fname, 'admin')
            QtWidgets.QMessageBox.information(self, "Success", "Admin account created successfully.\nStarting PekoCMS...")
            self.accept()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Error", f"Could not create user: {str(e)}")

    def restore_backup(self):
        zip_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Select Backup Archive", "", "ZIP Files (*.zip)")
        if zip_path:
            from app.restore_utils import restore_system_from_backup
            success = restore_system_from_backup(zip_path, self)
            if success:
                # Close app to allow restart
                QtWidgets.QApplication.quit()
                sys.exit(0)


class MainWindow(QtWidgets.QMainWindow):
    logout_signal = QtCore.Signal()
    
    # Button styling constants
    STYLE_PRIMARY_BTN = 'font-size: 12px; padding: 8px 24px; background-color: #0078D4; color: white; font-weight: bold;'
    STYLE_SECONDARY_BTN = 'font-size: 11px; padding: 8px 16px; background-color: #0078D4; color: white; font-weight: bold;'
    STYLE_SUCCESS_BTN = 'font-size: 11px; padding: 8px 16px; background-color: #4CAF50; color: white; font-weight: bold;'
    STYLE_DANGER_BTN = 'font-size: 12px; padding: 8px 24px; background-color: #D32F2F; color: white; font-weight: bold;'
    STYLE_SMALL_BTN = 'font-size: 10px; padding: 4px 8px;'
    STYLE_TINY_BTN = 'font-size: 11px; padding: 4px;'
    
    @staticmethod
    def style_button_with_dynamic_spacing(button, font_size=12, padding="10px 20px"):
        """Apply consistent button styling with dynamic sizing"""
        button.setStyleSheet(f'font-weight: bold; padding: {padding}; font-size: {font_size}px;')
        button.setSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Fixed)
    
    def __init__(self, user):
        super().__init__()
        # Initialize databases
        catalogue_db.init_db()
        polyclinic_db.init_db()
        
        self.user = user
        self.setWindowTitle(APP_NAME)
        self.resize(1200, 800)
        self.is_shutting_down = False
        self.current_mode = 'pathology'  # Start in pathology mode
        
        # Initialize polyclinic data storage
        self.polyclinic_data = {'doctors': [], 'bookings': []}
        
        # Smart Layout Detection
        screen = QtWidgets.QApplication.primaryScreen()
        screen_geo = screen.availableGeometry()
        # If height is small (e.g. 768p), enable compact mode
        self.compact_mode = screen_geo.height() < 900
        print(f"Screen Height: {screen_geo.height()}px -> Compact Mode: {self.compact_mode}")
        
        # Create central widget with header + tabs layout
        central = QtWidgets.QWidget()
        main_layout = QtWidgets.QVBoxLayout(central)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        # Header bar with neutral/system themed branding
        header = QtWidgets.QWidget()
        header.setObjectName('lightHeader')
        header_layout = QtWidgets.QHBoxLayout(header)
        header_layout.setContentsMargins(10, 8, 10, 8)
        
        # Logo
        logo = QtWidgets.QLabel()
        pix = render_svg(get_asset_path(LOGO_SVG), 40) # Crisp render at 40px
        if pix:
            logo.setPixmap(pix)
        header_layout.addWidget(logo)
        
        # Clinic name
        clinic_name = QtWidgets.QLabel(APP_NAME)
        clinic_name.setStyleSheet('font-size: 22px; font-weight: bold; margin-left: 10px;')
        header_layout.addWidget(clinic_name)
        
        # Mode toggle buttons
        header_layout.addSpacing(20)
        self.pathology_btn = QtWidgets.QPushButton(MODE_PATHOLOGY)
        self.pathology_btn.setStyleSheet('font-size: 12px; font-weight: bold; padding: 6px 16px; background-color: #0078D4; color: white; border-radius: 4px;')
        self.pathology_btn.clicked.connect(lambda: self.switch_mode('pathology'))
        header_layout.addWidget(self.pathology_btn)
        
        self.polyclinic_btn = QtWidgets.QPushButton(MODE_POLYCLINIC)
        self.polyclinic_btn.setStyleSheet('font-size: 12px; font-weight: bold; padding: 6px 16px; background-color: #A9A9A9; color: white; border-radius: 4px;')
        self.polyclinic_btn.clicked.connect(lambda: self.switch_mode('polyclinic'))
        header_layout.addWidget(self.polyclinic_btn)
        
        # UI Scale control
        header_layout.addSpacing(20)
        scale_label = QtWidgets.QLabel('Scale:')
        scale_label.setStyleSheet('font-size: 11px; font-weight: bold;')
        header_layout.addWidget(scale_label)
        
        self.scale_slider = QtWidgets.QSlider(QtCore.Qt.Horizontal)
        self.scale_slider.setMinimum(50)  # 50%
        self.scale_slider.setMaximum(150)  # 150%
        self.scale_slider.setValue(int(UI_SCALE * 100))  # Load from config
        self.scale_slider.setTickPosition(QtWidgets.QSlider.TicksBelow)
        self.scale_slider.setTickInterval(25)
        self.scale_slider.setMaximumWidth(120)
        self.scale_slider.setStyleSheet('QSlider { margin: 0 5px; }')
        self.scale_slider.valueChanged.connect(self.on_scale_changed)
        header_layout.addWidget(self.scale_slider)
        
        self.scale_value_label = QtWidgets.QLabel(f'{int(UI_SCALE * 100)}%')
        self.scale_value_label.setStyleSheet('font-size: 11px; font-weight: bold; min-width: 40px;')
        header_layout.addWidget(self.scale_value_label)
        
        # Compact Mode Toggle
        header_layout.addSpacing(10)
        self.compact_check = QtWidgets.QCheckBox("Compact")
        self.compact_check.setStyleSheet('font-weight: bold; font-size: 11px;')
        self.compact_check.setChecked(self.compact_mode)
        self.compact_check.toggled.connect(self.on_compact_toggled)
        header_layout.addWidget(self.compact_check)
        
        # User info and logout button
        header_layout.addStretch()
        user_label = QtWidgets.QLabel(f"User: {user.get('username', 'Unknown')}")
        user_label.setStyleSheet('font-size: 13px;')
        header_layout.addWidget(user_label)
        
        logout_btn = QtWidgets.QPushButton('Logout/Login')
        logout_btn.setStyleSheet('font-size: 11px; padding: 4px 8px; background-color: #f44336; color: white;')
        logout_btn.clicked.connect(self.do_logout)
        header_layout.addWidget(logout_btn)
        
        shutdown_btn = QtWidgets.QPushButton('Shutdown')
        shutdown_btn.setStyleSheet('font-size: 11px; padding: 4px 8px; background-color: #d32f2f; color: white;')
        shutdown_btn.clicked.connect(self.do_shutdown)
        header_layout.addWidget(shutdown_btn)
        
        main_layout.addWidget(header)
        
        # Create stacked widget for mode switching
        self.mode_stack = QtWidgets.QStackedWidget()
        
        # ===== PATHOLOGY MODE =====
        pathology_widget = QtWidgets.QWidget()
        pathology_layout = QtWidgets.QVBoxLayout(pathology_widget)
        self.pathology_tabs = QtWidgets.QTabWidget()
        pathology_layout.addWidget(self.pathology_tabs)
        
        # Create pathology tab widgets
        self.invoice_tab = QtWidgets.QWidget()
        self.cms_tab = QtWidgets.QWidget()
        self.datasheet_tab = QtWidgets.QWidget()
        self.reports_tab = QtWidgets.QWidget()
        
        self.pathology_tabs.addTab(self.invoice_tab, "Invoice Generator")
        self.pathology_tabs.addTab(self.cms_tab, "Patient CMS")
        self.pathology_tabs.addTab(self.datasheet_tab, "Datasheet")
        self.pathology_tabs.addTab(self.reports_tab, "Report Tracker")
        
        if user.get('role') == 'admin':
            self.admin_tab = QtWidgets.QWidget()
            self.pathology_tabs.addTab(self.admin_tab, "Admin")
        
        self.mode_stack.addWidget(pathology_widget)
        
        # ===== POLYCLINIC MODE (Placeholder - will implement next) =====
        polyclinic_widget = QtWidgets.QWidget()
        polyclinic_layout = QtWidgets.QVBoxLayout(polyclinic_widget)
        self.polyclinic_tabs = QtWidgets.QTabWidget()
        polyclinic_layout.addWidget(self.polyclinic_tabs)
        
        # Create polyclinic tab widgets (placeholders for now)
        self.poly_booking_tab = QtWidgets.QWidget()
        self.poly_doctor_tab = QtWidgets.QWidget()
        self.poly_queue_tab = QtWidgets.QWidget()
        self.poly_cms_tab = QtWidgets.QWidget()
        
        self.polyclinic_tabs.addTab(self.poly_booking_tab, "Patient Booking")
        self.polyclinic_tabs.addTab(self.poly_doctor_tab, "Doctor Entry/Lookup")
        self.polyclinic_tabs.addTab(self.poly_queue_tab, "Patient Queue View")
        self.polyclinic_tabs.addTab(self.poly_cms_tab, "Patient CMS")
        
        self.mode_stack.addWidget(polyclinic_widget)
        
        self.mode_stack.addWidget(polyclinic_widget)
        
        main_layout.addWidget(self.mode_stack)
        
        self.setCentralWidget(central)
        
        # Apply initial scale from config
        # Smart Auto-Scale: If we are in compact mode (small screen) and user hasn't customized scale (still 1.0),
        # automatically drop to 0.85 which is the sweet spot for 768p
        current_s = UI_SCALE
        if self.compact_mode and current_s == 1.0:
            print("Auto-scaling to 0.85 for 768p screen")
            current_s = 0.85
            # We don't save this to config automatically to avoid overriding user preference persistently if they move screens,
            # but we update the UI
            
        self.apply_ui_scale(current_s)
        self.scale_slider.setValue(int(current_s * 100)) # Sync slider
        
        # Initialize pathology tabs
        self.init_invoice_tab()
        self.init_cms_tab()
        self.init_datasheet_tab()
        self.init_reports_tab()
        if user.get('role') == 'admin':
            self.init_admin_tab()
        
        # Initialize polyclinic tabs
        self.init_poly_booking_tab()
        self.init_poly_doctor_tab()
        self.init_poly_queue_tab()
        self.init_poly_cms_tab()
        
        # Add footer
        footer = QtWidgets.QLabel(FOOTER_TEXT)
        footer.setAlignment(QtCore.Qt.AlignCenter)
        footer.setStyleSheet('color: gray; font-size: 10px; padding: 4px;')
        main_layout.addWidget(footer)
        
        # Refresh database on login
        self.refresh_database()

        # Check for updates
        QtCore.QTimer.singleShot(1000, lambda: check_for_updates_gui(self))

        # Connect tab change signals for auto-resizing
        self.pathology_tabs.currentChanged.connect(self.on_tab_changed)
        self.polyclinic_tabs.currentChanged.connect(self.on_tab_changed)

        # Preload Special Tests
        self.preload_special_tests()
    
    def on_scale_changed(self, value):
        """Handle scale slider changes"""
        scale = value / 100.0
        self.scale_value_label.setText(f'{value}%')
        self.apply_ui_scale(scale)
        self.save_scale_to_config(scale)

    def on_compact_toggled(self, checked):
        """Handle compact mode toggle"""
        self.compact_mode = checked
        print(f"Compact Mode toggled: {self.compact_mode}")
        
        # Rebuild Invoice Tab
        # 1. Clear existing layout safely
        if self.invoice_tab.layout():
            self.clear_layout(self.invoice_tab.layout())
            QtWidgets.QWidget().setLayout(self.invoice_tab.layout()) # Hack to unparent layout if needed, but deleteLater usually works
            # Actually, just deleting the old layout widget items is enough?
            # layout remains set on the widget. We need to repurpose it or delete it.
            # self.invoice_tab.layout().deleteLater() # This doesn't remove it immediately
            
            # Better approach: Create a new widget for the tab to ensure clean slate?
            # But the tab widget holds the reference.
            
            # Let's use the clear_layout helper and reuse the existing layout instance if possible?
            # No, standard practice is to delete the layout item.
            old_layout = self.invoice_tab.layout()
            self.clear_layout(old_layout)
            # Remove layout from widget
            # self.invoice_tab.setLayout(None) # Not supported in PySide
            
            # Just create a new container widget for the tab content!
            # Wait, self.invoice_tab IS the widget in the tab.
            # Let's just create a new widget and replace the tab.
            
            new_tab = QtWidgets.QWidget()
            # Replace in QTabWidget
            idx = self.pathology_tabs.indexOf(self.invoice_tab)
            if idx >= 0:
                self.pathology_tabs.removeTab(idx)
                self.invoice_tab = new_tab # Update reference
                self.pathology_tabs.insertTab(idx, self.invoice_tab, "Invoice Generator")
                self.pathology_tabs.setCurrentIndex(idx) # Keep focus
                
                # Re-init
                self.init_invoice_tab()
    
    def clear_layout(self, layout):
        """Recursively clear a layout"""
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()
                else:
                    self.clear_layout(item.layout())
            layout.deleteLater()
    
    class ResizingGraphicsView(QtWidgets.QGraphicsView):
        def __init__(self, scene, parent=None):
            super().__init__(scene, parent)
            self.setFrameShape(QtWidgets.QFrame.NoFrame)
            self.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
            self.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
            self.setAlignment(QtCore.Qt.AlignTop | QtCore.Qt.AlignLeft)
            self.setRenderHint(QtGui.QPainter.Antialiasing)
            self.setRenderHint(QtGui.QPainter.SmoothPixmapTransform)
            self.main_window = None # Reference to main window to access scale factor

        def resizeEvent(self, event):
            super().resizeEvent(event)
            if self.scene() and self.scene().items():
                # Find the proxy widget (should be the first item)
                proxy = self.scene().items()[0]
                if isinstance(proxy, QtWidgets.QGraphicsProxyWidget):
                    scale = getattr(self.main_window, 'current_scale', 1.0) if self.main_window else 1.0
                    
                    # Target: We want the widget (scaled) to fit the viewport exactly.
                    # Visual Size = Viewport Size
                    # Logical Widget Size = Viewport Size / Scale
                    
                    vw = self.viewport().width()
                    vh = self.viewport().height()
                    
                    new_width = vw / scale
                    new_height = vh / scale
                    
                    # Update widget size (Logical)
                    if proxy.widget():
                        proxy.widget().setFixedSize(int(new_width), int(new_height))
                        
                    # Update scene rect (Visual)
                    # The scene rect should match the viewport size because the Item's transform 
                    # shrinks it to that visual size.
                    self.scene().setSceneRect(0, 0, vw, vh)

    def apply_ui_scale(self, scale):
        """Apply UI scale transformation to central widget using QGraphicsView"""
        try:
            # Store the scale for reference
            self.current_scale = scale
            
            # Get the current central widget
            current_central = self.centralWidget()
            if not current_central:
                return
            
            # Check if we are already in scaled mode (current_central is our wrapper)
            is_already_scaled = isinstance(current_central, MainWindow.ResizingGraphicsView)
            
            if scale == 1.0:
                if is_already_scaled:
                    # Unwrap: Retrieve the original widget
                    proxy = current_central.scene().items()[-1] # Usually items are stacked, check properly
                    # Better: usage referencing existing stored widget
                    if hasattr(self, '_original_central_widget') and self._original_central_widget:
                        try:
                            # Restore original - Explicitly reparent to self to ensure ownership is claimed
                            self.setCentralWidget(self._original_central_widget)
                            # Reset fixed size constraints by setting to default min/max
                            self._original_central_widget.setMinimumSize(0, 0)
                            self._original_central_widget.setMaximumSize(16777215, 16777215)
                            self._original_central_widget.updateGeometry()
                            
                            # Clean up the graphics view infrastructure
                            current_central.deleteLater()
                            if hasattr(self, 'graphics_scene'):
                                 self.graphics_scene.deleteLater()
                                 del self.graphics_scene
                        except RuntimeError:
                            # Widget might be deleted during shutdown or scaling update
                            pass
                        
                        if hasattr(self, 'proxy_widget'):
                            del self.proxy_widget
                            
                        del self._original_central_widget
            else:
                # We want to scale
                if not is_already_scaled:
                    # FIRST: Detach the original central widget from MainWindow so it can be embedded
                    # takeCentralWidget passes ownership to us
                    self._original_central_widget = self.takeCentralWidget()
                    
                    if not self._original_central_widget:
                        # Should not happen if we checked centralWidget() before, but safety first
                        # Attempt to recover by using the reference we got earlier, assuming we can reparent it
                        # But if takeCentralWidget returned None, we might be in trouble.
                        self._original_central_widget = current_central
    
                    # Create graphics view setup
                    self.graphics_scene = QtWidgets.QGraphicsScene()
                    self.graphics_view = self.ResizingGraphicsView(self.graphics_scene, self)
                    self.graphics_view.main_window = self
                    
                    # Retrieve background color from original widget or window to fill gaps
                    bg_color = self.palette().color(QtGui.QPalette.Window)
                    self.graphics_view.setBackgroundBrush(bg_color)
                    
                    # Add the original widget as a proxy
                    self.proxy_widget = self.graphics_scene.addWidget(self._original_central_widget)
                    
                    # Set view as central
                    self.setCentralWidget(self.graphics_view)
                else:
                    # Already scaled, just updating scale factor
                    self.graphics_view = current_central
                    # Find proxy
                    for item in self.graphics_scene.items():
                        if isinstance(item, QtWidgets.QGraphicsProxyWidget):
                            self.proxy_widget = item
                            break
                
                # Apply transformation
                transform = QtGui.QTransform()
                transform.scale(scale, scale)
                self.proxy_widget.setTransform(transform)
                
                # Force a resize event to layout correctly with new scale
                # We manually trigger the logic from resizeEvent
                view_width = self.graphics_view.viewport().width()
                view_height = self.graphics_view.viewport().height()
                
                if view_width > 0 and view_height > 0:
                    new_width = view_width / scale
                    new_height = view_height / scale
                    self._original_central_widget.setFixedSize(int(new_width), int(new_height))
                    # VISUAL size matches viewport
                    self.graphics_scene.setSceneRect(0, 0, view_width, view_height)
        except Exception as e:
            print(f"Error applying UI scale: {e}")
            import traceback
            traceback.print_exc()
    
    def save_scale_to_config(self, scale):
        """Save the current scale to config.yaml"""
        try:
            config_path = get_config_path()
            
            # Load existing config
            config = {}
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = yaml.safe_load(f) or {}
            
            # Update scale
            config['UI_SCALE'] = scale
            
            # Save back to file
            with open(config_path, 'w', encoding='utf-8') as f:
                yaml.dump(config, f, default_flow_style=False, allow_unicode=True)
        except Exception as e:
            print(f"Warning: Could not save UI scale to config: {e}")

    def preload_special_tests(self):
        """Load special tests into memory on startup"""
        try:
            print("Preloading Special Tests...")
            self.special_tests_cache = special_tests_db.get_all_special_tests()
            print(f"Loaded {len(self.special_tests_cache)} special tests into memory.")
        except Exception as e:
            print(f"Error preloading special tests: {e}")

    def on_tab_changed(self, index):
        """Auto-resize window when switching tabs"""
        # Process events to ensure layout is updated before resizing
        QtWidgets.QApplication.processEvents()
        # self.adjustSize() - Disabled per user request
    
    def switch_mode(self, mode: str):
        """Switch between Pathology and Polyclinic modes"""
        self.current_mode = mode
        if mode == 'pathology':
            self.mode_stack.setCurrentIndex(0)
            self.pathology_btn.setStyleSheet('font-size: 12px; font-weight: bold; padding: 6px 16px; background-color: #0078D4; color: white; border-radius: 4px;')
            self.polyclinic_btn.setStyleSheet('font-size: 12px; font-weight: bold; padding: 6px 16px; background-color: #A9A9A9; color: white; border-radius: 4px;')
        else:
            self.mode_stack.setCurrentIndex(1)
            self.pathology_btn.setStyleSheet('font-size: 12px; font-weight: bold; padding: 6px 16px; background-color: #A9A9A9; color: white; border-radius: 4px;')
            self.polyclinic_btn.setStyleSheet('font-size: 12px; font-weight: bold; padding: 6px 16px; background-color: #0078D4; color: white; border-radius: 4px;')
    
    def closeEvent(self, event):
        """Handle window close properly"""
        if self.is_shutting_down:
            # Allow the window to close if we're shutting down
            event.accept()
        else:
            # Otherwise, just logout (don't close the window)
            self.do_logout()
            event.ignore()
    
    def do_logout(self):
        """Emit logout signal to switch to login screen"""
        self.logout_signal.emit()
        self.hide()
    
    def do_shutdown(self):
        """Shutdown the application"""
        reply = QtWidgets.QMessageBox.question(
            self,
            'Confirm Shutdown',
            'Are you sure you want to shut down the application?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        if reply == QtWidgets.QMessageBox.Yes:
            self.is_shutting_down = True
            QtWidgets.QApplication.quit()
    
    def refresh_database(self):
        """Refresh the test catalogue database using QThread"""
        # Create and start catalogue loader thread
        self.catalogue_thread = CatalogueLoaderThread()
        self.catalogue_thread.status_updated.connect(self._update_catalogue_status)
        self.catalogue_thread.finished.connect(self._on_catalogue_loaded)
        self.catalogue_thread.start()
        
        # Load doctor list for polyclinic
        self._poly_load_doctors_once()
    
    def _on_catalogue_loaded(self):
        """Called when catalogue loading is complete"""
        # Clean up thread reference
        if hasattr(self, 'catalogue_thread'):
            self.catalogue_thread.quit()
            self.catalogue_thread.wait()
    
    def _on_polyclinic_data_loaded(self, data):
        """Called when polyclinic data loading is complete"""
        self.polyclinic_data = data
        # Clean up thread reference
        if hasattr(self, 'polyclinic_thread'):
            self.polyclinic_thread.quit()
            self.polyclinic_thread.wait()
    
    def _update_catalogue_status(self, status_msg):
        """Update catalogue status label"""
        try:
            if hasattr(self, 'cat_status'):
                self.cat_status.setText(status_msg)
        except RuntimeError:
            # Widget might be deleted during shutdown or scaling update
            pass
    
    def refresh_invoice_catalogue(self):
        
        # Load in background using QThread
        self.invoice_refresh_thread = InvoiceCatalogueLoaderThread()
        self.invoice_refresh_thread.catalogue_loaded.connect(self._on_invoice_catalogue_refreshed)
        self.invoice_refresh_thread.error_occurred.connect(lambda msg: self._on_refresh_error(msg))
        self.invoice_refresh_thread.start()
    
    def _on_refresh_timeout(self):
        """Handle refresh timeout to prevent freeze"""
        if hasattr(self, 'progress_dialog') and self.progress_dialog.isVisible():
            self.progress_dialog.close()
            QtWidgets.QMessageBox.warning(self, 'Timeout', 'Catalogue loading timed out. Please try again.')
    
    def _on_refresh_error(self, msg):
        """Handle refresh error"""
        if hasattr(self, 'progress_dialog'):
            self.progress_dialog.close()
        QtWidgets.QMessageBox.warning(self, 'Error', msg)
    
    def _on_invoice_catalogue_refreshed(self, cat, status_msg):
        """Callback when catalogue refresh is complete"""
        # Close progress dialog
        if hasattr(self, 'progress_dialog'):
            self.progress_dialog.close()
        
        # Catalogue is now cached and ready for search
        try:
            if hasattr(self, 'cat_status'):
                self.cat_status.setText(f'Ready - {len(cat) if cat else 0} tests cached')
            
            QtWidgets.QMessageBox.information(self, 'Success', f'Catalogue refreshed: {len(cat)} tests loaded')
        except RuntimeError:
            pass
        
        # Clean up thread
        if hasattr(self, 'invoice_refresh_thread'):
            self.invoice_refresh_thread.quit()
            self.invoice_refresh_thread.wait()
    
    def refresh_full_catalogue(self):
        """Refresh the full catalogue tab using background thread"""
        # Load catalogue data in background
        self.full_cat_thread = FullCatalogueLoaderThread()
        self.full_cat_thread.data_ready.connect(self._on_full_catalogue_data_ready)
        self.full_cat_thread.error_occurred.connect(lambda msg: QtWidgets.QMessageBox.warning(self, 'Error', msg))
        self.full_cat_thread.start()
    
    def _on_full_catalogue_data_ready(self, data):
        """Render full catalogue table after data is loaded"""
        try:
            if hasattr(self, 'cat_full_table'):
                q = ''
                if hasattr(self, 'cat_search_widget'):
                    q = self.cat_search_widget.text().lower()
                
                # Batch rendering
                self.cat_full_table.setUpdatesEnabled(False)
                self.cat_full_table.setRowCount(0)
                
                try:
                    row_count = 0
                    for t in data:
                        if q in (t.get('testCode', '') + t.get('testName', '')).lower():
                            self.cat_full_table.insertRow(row_count)
                            self.cat_full_table.setItem(row_count, 0, QtWidgets.QTableWidgetItem(t.get('testCode', '')))
                            self.cat_full_table.setItem(row_count, 1, QtWidgets.QTableWidgetItem(t.get('testName', '')))
                            self.cat_full_table.setItem(row_count, 2, QtWidgets.QTableWidgetItem(f"₹{t.get('testFees', 0):.2f}"))
                            self.cat_full_table.setItem(row_count, 3, QtWidgets.QTableWidgetItem(t.get('CategoryName', '')))
                            row_count += 1
                finally:
                    self.cat_full_table.setUpdatesEnabled(True)
                    self.cat_full_table.resizeRowsToContents()
                    self.cat_full_table.update()
                
                QtWidgets.QMessageBox.information(self, 'Success', 'Catalogue refreshed successfully')
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Error', f'Failed to refresh catalogue: {str(e)}')
        finally:
            # Clean up thread
            if hasattr(self, 'full_cat_thread'):
                self.full_cat_thread.quit()
                self.full_cat_thread.wait()
    
    
    def poly_refresh_booking_doctors(self):
        """Refresh the doctor list in patient booking tab"""
        try:
            self.poly_doctors_loaded = False
            self._poly_load_doctors_once()
            QtWidgets.QMessageBox.information(self, 'Success', 'Doctor list refreshed')
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Error', f'Failed to refresh doctors: {str(e)}')
    
    def poly_refresh_doctor_list(self):
        """Refresh the doctor list in doctor entry/lookup tab"""
        try:
            self.poly_reload_doctor_list()
            QtWidgets.QMessageBox.information(self, 'Success', 'Doctor list refreshed successfully')
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Error', f'Failed to refresh doctor list: {str(e)}')
    
    def _poly_populate_doctor_mgmt_table(self, doctors):
        """Populate doctor management table - deprecated, use poly_reload_doctor_list instead"""
        pass  # This is handled by poly_reload_doctor_list()
    
    # ===== INVOICE TAB ====
    def init_invoice_tab(self):
        layout = QtWidgets.QVBoxLayout(self.invoice_tab)
        
        # ADAPTIVE LAYOUT SETTINGS
        if self.compact_mode:
            # Compact Mode (for 768p screens)
            margin = 8
            spacing = 6
            cat_stretch = 3 # More space for available tests (Spotlight)
            sel_stretch = 1 # Less space for selected tests (Compact)
            table_min_h = 120
            sel_table_min_h = 100
            input_font_size = 10
        else:
            # Standard Mode (for 1080p+ screens)
            margin = 15
            spacing = 12
            cat_stretch = 3
            sel_stretch = 1
            table_min_h = 150
            sel_table_min_h = 120
            input_font_size = 11

        layout.setContentsMargins(margin, margin, margin, margin)
        layout.setSpacing(spacing)
        
        # ===== TOP: CATALOGUE SECTION =====
        cat_group = QtWidgets.QGroupBox('Available Tests')
        cat_group.setObjectName('availableTestsGroup')
        cat_layout = QtWidgets.QVBoxLayout(cat_group)
        # Compact group box changes
        if self.compact_mode:
            cat_layout.setContentsMargins(4, 12, 4, 4) # Tighter inside groupbox
            cat_layout.setSpacing(4)
        
        # Search bar with button
        search_layout = QtWidgets.QHBoxLayout()
        search_label = QtWidgets.QLabel('Search:')
        search_label.setFont(QtGui.QFont('', input_font_size))
        search_layout.addWidget(search_label)
        self.cat_search = QtWidgets.QLineEdit()
        self.cat_search.setPlaceholderText('By test name or code...')
        self.cat_search.setFont(QtGui.QFont('', input_font_size))
        self.cat_search.returnPressed.connect(self.search_invoice_catalogue)
        search_layout.addWidget(self.cat_search)
        
        # Search button (on-demand instead of auto-filter)
        self.cat_search_btn = QtWidgets.QPushButton('🔍 Search')
        if self.compact_mode:
            self.cat_search_btn.setStyleSheet('font-size: 11px; padding: 4px 8px; font-weight: bold;')
        else:
            self.cat_search_btn.setStyleSheet('font-size: 14px; padding: 6px 12px; font-weight: bold;')
            
        self.cat_search_btn.clicked.connect(self.search_invoice_catalogue)
        self.cat_search_btn.setMinimumWidth(80 if self.compact_mode else 100)
        search_layout.addWidget(self.cat_search_btn)
        
        self.cat_status = QtWidgets.QLabel('Ready')
        self.cat_status.setStyleSheet('font-size: 10px; margin-left: 10px;')
        search_layout.addWidget(self.cat_status)
        
        # Refresh button for catalogue
        self.cat_refresh_btn = QtWidgets.QPushButton('🔄 Refresh')
        if self.compact_mode:
            self.cat_refresh_btn.setStyleSheet('font-size: 11px; padding: 4px 8px; font-weight: bold;')
        else:
            self.cat_refresh_btn.setStyleSheet('font-size: 14px; padding: 6px 12px; font-weight: bold;')
            
        self.cat_refresh_btn.clicked.connect(self.refresh_invoice_catalogue)
        self.cat_refresh_btn.setMinimumWidth(80 if self.compact_mode else 100)
        search_layout.addWidget(self.cat_refresh_btn)
        
        cat_layout.addLayout(search_layout)
        
        # Tabs for Standard and Special Tests
        self.cat_tabs = QtWidgets.QTabWidget()
        
        # Standard Tests tab
        standard_tab = QtWidgets.QWidget()
        standard_layout = QtWidgets.QVBoxLayout(standard_tab)
        if self.compact_mode:
            standard_layout.setContentsMargins(4, 4, 4, 4)
            standard_layout.setSpacing(4)
            
        self.cat_table = QtWidgets.QTableWidget()
        self.cat_table.setColumnCount(5)
        self.cat_table.setHorizontalHeaderLabels(['Code', 'Name', 'Fasting', 'Fees', 'Action'])
        self.cat_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.cat_table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.cat_table.setMinimumHeight(table_min_h) 
        self.cat_table.setAlternatingRowColors(True)
        font = self.cat_table.font()
        font.setPointSize(input_font_size)
        self.cat_table.setFont(font)
        self.cat_table.setColumnWidth(2, 70)
        self.cat_table.setColumnWidth(4, 80)
        self.cat_table.setStyleSheet("""
            QTableWidget { alternate-background-color: palette(base); background-color: palette(base); }
            QTableWidget::item:selected { background-color: palette(highlight); color: palette(highlighted-text); }
        """)
        standard_layout.addWidget(self.cat_table)
        self.cat_tabs.addTab(standard_tab, "Standard Tests")
        
        # Special Tests tab
        special_tab = QtWidgets.QWidget()
        special_layout = QtWidgets.QVBoxLayout(special_tab)
        special_layout.setSpacing(spacing)
        
        # Toggle button for adding new special tests
        self.toggle_add_test_btn = QtWidgets.QPushButton('▶ Add New Special Test')
        MainWindow.style_button_with_dynamic_spacing(self.toggle_add_test_btn, font_size=input_font_size, padding="4px 8px" if self.compact_mode else "6px 12px")
        self.toggle_add_test_btn.setCheckable(True)
        self.toggle_add_test_btn.setStyleSheet("text-align: left; font-weight: bold;")
        special_layout.addWidget(self.toggle_add_test_btn)
        
        # Form to add new special tests (initially hidden)
        self.add_test_form_widget = QtWidgets.QWidget()
        self.add_test_form_widget.setVisible(False)
        form_layout = QtWidgets.QVBoxLayout(self.add_test_form_widget)
        form_layout.setContentsMargins(0, 0, 0, 0)
        
        self.toggle_add_test_btn.toggled.connect(lambda checked: self.add_test_form_widget.setVisible(checked))
        self.toggle_add_test_btn.toggled.connect(lambda checked: self.toggle_add_test_btn.setText('▼ Add New Special Test' if checked else '▶ Add New Special Test'))
        
        # Only show the form if user is admin
        is_admin = self.user.get('role') == 'admin'
        # Only show the toggle if user is admin
        is_admin = self.user.get('role') == 'admin'
        self.toggle_add_test_btn.setVisible(is_admin)
        
        # Test Name
        name_layout = QtWidgets.QHBoxLayout()
        name_layout.addWidget(QtWidgets.QLabel('Test Name:'))
        self.special_test_name = QtWidgets.QLineEdit()
        self.special_test_name.setPlaceholderText('Enter test name')
        name_layout.addWidget(self.special_test_name)
        form_layout.addLayout(name_layout)
        
        # Description
        desc_layout = QtWidgets.QHBoxLayout()
        desc_layout.addWidget(QtWidgets.QLabel('Description:'))
        self.special_test_desc = QtWidgets.QLineEdit()
        self.special_test_desc.setPlaceholderText('Enter test description')
        desc_layout.addWidget(self.special_test_desc)
        form_layout.addLayout(desc_layout)
        
        # Fees
        fees_layout = QtWidgets.QHBoxLayout()
        fees_layout.addWidget(QtWidgets.QLabel('Fees (₹):'))
        self.special_test_fees = QtWidgets.QDoubleSpinBox()
        self.special_test_fees.setMinimum(0)
        self.special_test_fees.setMaximum(100000)
        self.special_test_fees.setDecimals(2)
        self.special_test_fees.setValue(0)
        fees_layout.addWidget(self.special_test_fees)
        fees_layout.addStretch()
        form_layout.addLayout(fees_layout)
        
        # Add button
        self.special_test_add_btn = QtWidgets.QPushButton('Add Special Test')
        self.special_test_add_btn.clicked.connect(self.add_new_special_test)
        form_layout.addWidget(self.special_test_add_btn)
        
        special_layout.addWidget(self.add_test_form_widget)
        
        # Special tests table
        self.special_table = QtWidgets.QTableWidget()
        self.special_table.setColumnCount(4)
        self.special_table.setHorizontalHeaderLabels(['Name', 'Description', 'Fees', 'Action'])
        self.special_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.special_table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.special_table.setMinimumHeight(table_min_h)
        self.special_table.setAlternatingRowColors(True)
        font = self.special_table.font()
        font.setPointSize(input_font_size)
        self.special_table.setFont(font)
        self.special_table.setColumnWidth(3, 80)
        self.special_table.setStyleSheet("""
            QTableWidget { alternate-background-color: palette(base); background-color: palette(base); }
            QTableWidget::item:selected { background-color: palette(highlight); color: palette(highlighted-text); }
        """)
        special_layout.addWidget(self.special_table)
        self.cat_tabs.addTab(special_tab, "Special Tests")
        
        cat_layout.addWidget(self.cat_tabs)
        
        layout.addWidget(cat_group, cat_stretch)
        
        # ===== MIDDLE: PATIENT & SELECTED TESTS (side by side) =====
        middle = QtWidgets.QWidget()
        middle_layout = QtWidgets.QHBoxLayout(middle)
        middle_layout.setSpacing(10)
        if self.compact_mode:
            middle_layout.setContentsMargins(0, 0, 0, 0)
        
        # Left: Patient Details (heavily scaled down)
        patient_group = QtWidgets.QGroupBox('Patient Details')
        patient_layout = QtWidgets.QVBoxLayout(patient_group)
        patient_layout.setContentsMargins(8, 8, 8, 8)
        patient_layout.setSpacing(4)
        
        lookup_layout = QtWidgets.QHBoxLayout()
        lookup_layout.setContentsMargins(0, 0, 0, 0)
        lookup_layout.setSpacing(3)
        self.inv_patient_id = QtWidgets.QLineEdit()
        self.inv_patient_id.setPlaceholderText('ID or Phone...')
        self.inv_patient_id.setMaximumHeight(26)
        self.inv_patient_id.setStyleSheet('font-size: 10px;')
        self.inv_lookup_btn = QtWidgets.QPushButton('Lookup')
        self.inv_lookup_btn.setMaximumHeight(26)
        self.inv_lookup_btn.setMaximumWidth(65)
        self.inv_lookup_btn.setStyleSheet('font-size: 10px;')
        self.inv_lookup_btn.clicked.connect(self.inv_lookup_patient)
        lookup_layout.addWidget(self.inv_patient_id)
        lookup_layout.addWidget(self.inv_lookup_btn)
        patient_layout.addLayout(lookup_layout)
        
        self.inv_patient_info = QtWidgets.QLabel('Not Selected')
        self.inv_patient_info.setStyleSheet('font-size: 13px; font-weight: bold; line-height: 1.4; color: #333;')
        self.inv_patient_info.setWordWrap(True)
        patient_layout.addWidget(self.inv_patient_info)
        patient_layout.addStretch()
        
        middle_layout.addWidget(patient_group, 0)
        
        # Right: Selected Tests
        tests_group = QtWidgets.QGroupBox('Selected Tests')
        tests_layout = QtWidgets.QVBoxLayout(tests_group)
        if self.compact_mode:
            tests_layout.setContentsMargins(4, 12, 4, 4)
            
        self.inv_items_table = QtWidgets.QTableWidget()
        self.inv_items_table.setObjectName('yellowTable')
        self.inv_items_table.setColumnCount(3)
        self.inv_items_table.setHorizontalHeaderLabels(['Name', 'Fees', 'Action'])
        self.inv_items_table.setMinimumHeight(sel_table_min_h) 
        self.inv_items_table.setAlternatingRowColors(True)
        self.inv_items_table.horizontalHeader().setStretchLastSection(False)
        self.inv_items_table.setColumnWidth(2, 80)
        # Increase font size
        font = self.inv_items_table.font()
        font.setPointSize(input_font_size)
        self.inv_items_table.setFont(font)
        tests_layout.addWidget(self.inv_items_table)
        middle_layout.addWidget(tests_group, 2)
        
        layout.addWidget(middle, sel_stretch)
        
        # ===== BOTTOM: BILLING & GENERATE =====
        # ===== BOTTOM: BILLING & GENERATE =====
        billing_group = QtWidgets.QGroupBox('Billing')
        # Use Grid Layout for compact row
        billing_layout = QtWidgets.QGridLayout(billing_group)
        billing_layout.setContentsMargins(10, 8, 10, 8)
        billing_layout.setHorizontalSpacing(15)
        
        # --- Column 0 & 1: Inputs ---
        self.inv_discount = QtWidgets.QDoubleSpinBox()
        self.inv_discount.setMaximum(100)
        self.inv_discount.setSuffix('%')
        self.inv_discount.setFont(QtGui.QFont('', input_font_size))
        self.inv_discount.setMinimumWidth(110) # Expanded "by a lot"
        self.inv_discount.valueChanged.connect(self.inv_recalc)
        
        self.inv_home_check = QtWidgets.QCheckBox('Home Coll.')
        self.inv_home_check.setFont(QtGui.QFont('', input_font_size))
        self.inv_home_check.toggled.connect(self.inv_home_toggled)
        
        self.inv_home_fee = QtWidgets.QDoubleSpinBox()
        self.inv_home_fee.setMaximum(100000)
        self.inv_home_fee.setFont(QtGui.QFont('', input_font_size))
        self.inv_home_fee.setMinimumWidth(110) # Expanded "by a lot"
        self.inv_home_fee.setEnabled(False)
        self.inv_home_fee.valueChanged.connect(self.inv_recalc)
        
        # Row 0: Discount
        billing_layout.addWidget(QtWidgets.QLabel('Disc %:'), 0, 0)
        billing_layout.addWidget(self.inv_discount, 0, 1)
        
        # Row 1: Home Check
        billing_layout.addWidget(self.inv_home_check, 1, 0, 1, 2)
        
        # Row 2: Home Fee
        billing_layout.addWidget(QtWidgets.QLabel('Fee:'), 2, 0)
        billing_layout.addWidget(self.inv_home_fee, 2, 1)
        
        # --- Column 2: Summary (Vertical) ---
        summary_layout = QtWidgets.QVBoxLayout()
        summary_layout.setSpacing(2)
        self.inv_subtotal = QtWidgets.QLabel('Sub: ₹0')
        self.inv_subtotal.setStyleSheet('font-size: 11px; color: #555;')
        self.inv_disc_amt = QtWidgets.QLabel('Disc: -₹0')
        self.inv_disc_amt.setStyleSheet('font-size: 11px; color: #d32f2f;')
        self.inv_total = QtWidgets.QLabel('Total: ₹0')
        self.inv_total.setStyleSheet('font-weight: bold; font-size: 14px; color: #2e7d32;')
        
        summary_layout.addWidget(self.inv_subtotal)
        summary_layout.addWidget(self.inv_disc_amt)
        summary_layout.addWidget(self.inv_total)
        
        billing_layout.addLayout(summary_layout, 0, 2, 3, 1)
        
        # --- Column 3: Spacer ---
        billing_layout.setColumnStretch(3, 1)
        
        # --- Column 4: Generate Button ---
        self.inv_gen_btn = QtWidgets.QPushButton('Generate\nInvoice')
        # Square-ish styling
        self.inv_gen_btn.setStyleSheet("""
            QPushButton {
                background-color: #0078D4; 
                color: white; 
                font-size: 13px; 
                font-weight: bold;
                border-radius: 4px;
                padding: 4px;
                text-align: center;
            }
            QPushButton:hover { background-color: #0063b1; }
            QPushButton:disabled { background-color: #cccccc; }
        """)
        self.inv_gen_btn.setFixedSize(90, 80)
        self.inv_gen_btn.clicked.connect(self.inv_generate)
        self.inv_gen_btn.setEnabled(False)
        
        billing_layout.addWidget(self.inv_gen_btn, 0, 4, 3, 1)
        
        layout.addWidget(billing_group)
        
        self.inv_current_patient = None
        self.inv_selected_tests = {}
        
        # Cache will be updated via refresh button - no automatic loading
        self.cat_status.setText('Ready')
        # Connect to tab change event to ensure catalogue is cached
        self.pathology_tabs.currentChanged.connect(self._on_pathology_tab_changed)
        
        # Populate special tests table on initialization
        self.search_special_tests()
    
    def _on_pathology_tab_changed(self, tab_index):
        """Handle pathology tab changes - on-demand loading via search button"""
        # No automatic clearing - form clears after invoice generation instead
        pass
    
    def _on_invoice_catalogue_loaded(self, cat, status_msg):
        """Callback when invoice catalogue is cached"""
        self.cat_status.setText(f'Ready - {len(cat) if cat else 0} tests cached')
        # Clean up thread
        if hasattr(self, 'invoice_cat_thread'):
            self.invoice_cat_thread.quit()
            self.invoice_cat_thread.wait()
    
    def _load_invoice_catalogue_background(self):
        """Load catalogue in background thread"""
        try:
            cat = catalogue_db.get_all_tests()
            if not cat:
                # Fall back to data fetcher if database is empty
                cat = data_fetcher.get_catalogue_data()
                status = data_fetcher.get_fetch_status()
                status_msg = status[:50]
            else:
                status_msg = f"Loaded {len(cat)} tests from cache"
            
            # Update on main thread
            QtCore.QTimer.singleShot(0, lambda: self._update_inv_catalogue_on_main_thread(cat, status_msg))
        except Exception as e:
            QtCore.QTimer.singleShot(0, lambda: self.cat_status.setText(f"Error loading: {str(e)}"))
    
    def _on_special_tests_loaded(self, special_tests):
        """Callback when special tests are cached"""
        self.cat_status.setText(f'Ready - {len(special_tests) if special_tests else 0} special tests cached')
        # Clean up thread
        if hasattr(self, 'special_tests_thread'):
            self.special_tests_thread.quit()
            self.special_tests_thread.wait()
    
    def load_inv_catalogue(self):
        """Explicitly load catalogue (used for manual refresh)"""
        # Try to load from database first (faster, offline access)
        cat = catalogue_db.get_all_tests()
        if cat:
            self.cat_status.setText(f"Loaded {len(cat)} tests from cache")
        else:
            # Fall back to data fetcher if database is empty
            cat = data_fetcher.get_catalogue_data()
            status = data_fetcher.get_fetch_status()
            self.cat_status.setText(status[:50])
        
        self.inv_all_catalogue = cat
        self.inv_catalogue_loaded = True
        self.filter_inv_catalogue()
    
    def load_special_tests(self):
        """Load special/custom tests from database"""
        self.special_tests_data = special_tests_db.get_all_special_tests()
        self.special_tests_loaded = True
        self.filter_special_tests()
    
    def search_invoice_catalogue(self):
        """Query SQLite catalogue on-demand and display results"""
        q = self.cat_search.text().lower().strip()
        
        if not q:
            self.cat_status.setText('Please enter a search term')
            self.cat_table.setRowCount(0)
            self.special_table.setRowCount(0)
            return
        
        try:
            self.cat_status.setText('Searching...')
            
            # Query standard tests from SQLite
            results = catalogue_db.search_tests(q)
            
            # Display results with batch rendering
            self.cat_table.setUpdatesEnabled(False)
            self.cat_table.setRowCount(0)
            
            row_count = 0
            for t in results:
                self.cat_table.insertRow(row_count)
                
                # Create items with selectable but non-editable flags
                flags = QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled
                code_item = QtWidgets.QTableWidgetItem(t.get('testCode', ''))
                code_item.setFlags(flags)
                name_item = QtWidgets.QTableWidgetItem(t.get('testName', ''))
                name_item.setFlags(flags)
                
                # Fasting indicator
                is_fasting = t.get('FastingRequired', '').lower() == 'yes'
                fasting_item = QtWidgets.QTableWidgetItem('Yes' if is_fasting else 'No')
                fasting_item.setFlags(flags)
                if is_fasting:
                    fasting_item.setBackground(QtGui.QColor(255, 100, 100))  # Red background
                    fasting_item.setForeground(QtGui.QColor(255, 255, 255))  # White text
                
                fees_item = QtWidgets.QTableWidgetItem(f"₹{t.get('testFees', 0):.2f}")
                fees_item.setFlags(flags)
                
                # Add button
                add_btn = QtWidgets.QPushButton('Add')
                add_btn.setMaximumWidth(70)
                add_btn.setStyleSheet('font-size: 11px; padding: 4px;')
                add_btn.clicked.connect(lambda checked, code=t.get('testCode'): self.add_test_to_selection(code))
                
                self.cat_table.setItem(row_count, 0, code_item)
                self.cat_table.setItem(row_count, 1, name_item)
                self.cat_table.setItem(row_count, 2, fasting_item)
                self.cat_table.setItem(row_count, 3, fees_item)
                self.cat_table.setCellWidget(row_count, 4, add_btn)
                
                row_count += 1
            
            self.cat_table.setUpdatesEnabled(True)
            self.cat_table.resizeRowsToContents()
            self.cat_table.update()
            
            # Query special tests
            self.search_special_tests()
            
            self.cat_status.setText(f'Found {row_count} standard tests')
        except Exception as e:
            self.cat_status.setText(f'Search error: {str(e)}')
    
    def search_special_tests(self):
        """Query special tests from database on-demand"""
        q = self.cat_search.text().lower().strip()
        
        try:
            # Query special tests from SQLite (always show all if no query)
            results = special_tests_db.search_special_tests(q) if q else special_tests_db.get_all_special_tests()
            
            # Display results with batch rendering
            self.special_table.setUpdatesEnabled(False)
            self.special_table.setRowCount(0)
            
            row_count = 0
            for t in results:
                self.special_table.insertRow(row_count)
                
                flags = QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled
                name_item = QtWidgets.QTableWidgetItem(t.get('testName', ''))
                name_item.setFlags(flags)
                
                desc_item = QtWidgets.QTableWidgetItem(t.get('testDescription', ''))
                desc_item.setFlags(flags)
                
                fees_item = QtWidgets.QTableWidgetItem(f"₹{t.get('testFees', 0):.2f}")
                fees_item.setFlags(flags)
                
                # Add button
                add_btn = QtWidgets.QPushButton('Add')
                add_btn.setMaximumWidth(70)
                add_btn.setStyleSheet('font-size: 11px; padding: 4px;')
                test_id = t.get('id')
                add_btn.clicked.connect(lambda checked, tid=test_id: self.add_special_test_to_selection(tid))
                
                self.special_table.setItem(row_count, 0, name_item)
                self.special_table.setItem(row_count, 1, desc_item)
                self.special_table.setItem(row_count, 2, fees_item)
                self.special_table.setCellWidget(row_count, 3, add_btn)
                
                row_count += 1
            
            self.special_table.setUpdatesEnabled(True)
            self.special_table.resizeRowsToContents()
            self.special_table.update()
        except Exception as e:
            self.cat_status.setText(f'Special tests error: {str(e)}')
    
    def add_special_test_to_selection(self, test_id: int):
        """Add a special test to the selected tests"""
        test_data = special_tests_db.get_special_test(test_id)
        if not test_data:
            return
        
        code = test_data.get('testCode')
        if code not in self.inv_selected_tests:
            self.inv_selected_tests[code] = {
                'testCode': code,
                'testName': test_data.get('testName', ''),
                'testFees': test_data.get('testFees', 0),
                'testDescription': test_data.get('testDescription', ''),
                'isSpecial': True
            }
            self.update_selected_tests_display()
            self.inv_recalc()
            self.inv_gen_btn.setEnabled(self.inv_current_patient is not None and len(self.inv_selected_tests) > 0)
    
    def add_test_to_selection(self, test_code: str):
        """Add a standard test from catalogue to the selected tests"""
        # Query the test from database to get full details
        test_data = catalogue_db.get_test(test_code)
        
        if not test_data:
            return
        
        if test_code not in self.inv_selected_tests:
            self.inv_selected_tests[test_code] = {
                'testCode': test_code,
                'testName': test_data.get('testName', ''),
                'testFees': test_data.get('testFees', 0),
                'testDescription': test_data.get('testDescription', ''),
                'isSpecial': False
            }
            self.update_selected_tests_display()
            self.inv_recalc()
            self.inv_gen_btn.setEnabled(self.inv_current_patient is not None and len(self.inv_selected_tests) > 0)
    
    def add_new_special_test(self):
        """Add a new special test to the database"""
        name = self.special_test_name.text().strip()
        desc = self.special_test_desc.text().strip()
        fees = self.special_test_fees.value()
        
        if not name:
            QtWidgets.QMessageBox.warning(self, 'Input Error', 'Please enter a test name.')
            return
        
        if fees <= 0:
            QtWidgets.QMessageBox.warning(self, 'Input Error', 'Please enter a valid fee amount.')
            return
        
        try:
            special_tests_db.add_special_test({
                'testName': name,
                'testDescription': desc,
                'testFees': fees
            })
            
            # Clear the form
            self.special_test_name.clear()
            self.special_test_desc.clear()
            self.special_test_fees.setValue(0)
            
            # Refresh the special tests list by re-searching
            self.search_special_tests()
            
            QtWidgets.QMessageBox.information(self, 'Success', f'Special test "{name}" added successfully!')
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, 'Error', f'Failed to add special test: {str(e)}')
    
    def update_selected_tests_display(self):
        """Refresh the selected tests table with remove buttons"""
        # Batch rendering
        self.inv_items_table.setUpdatesEnabled(False)
        self.inv_items_table.setRowCount(0)
        
        try:
            row_count = 0
            for code, test in self.inv_selected_tests.items():
                self.inv_items_table.insertRow(row_count)
                
                # Test name
                name_item = QtWidgets.QTableWidgetItem(test['testName'])
                name_item.setFlags(name_item.flags() & ~QtCore.Qt.ItemIsEditable)
                
                # Fees
                fees_item = QtWidgets.QTableWidgetItem(f"₹{test['testFees']:.2f}")
                fees_item.setFlags(fees_item.flags() & ~QtCore.Qt.ItemIsEditable)
                
                # Remove button
                remove_btn = QtWidgets.QPushButton('Remove')
                remove_btn.setMaximumWidth(70)
                remove_btn.setStyleSheet('font-size: 11px; padding: 4px;')
                remove_btn.clicked.connect(lambda checked, c=code: self.remove_selected_test(c))
                
                self.inv_items_table.setItem(row_count, 0, name_item)
                self.inv_items_table.setItem(row_count, 1, fees_item)
                self.inv_items_table.setCellWidget(row_count, 2, remove_btn)
                
                row_count += 1
        finally:
            self.inv_items_table.setUpdatesEnabled(True)
            self.inv_items_table.resizeRowsToContents()
            self.inv_items_table.update()
    
    def remove_selected_test(self, code: str):
        """Remove a test from selected tests"""
        if code in self.inv_selected_tests:
            del self.inv_selected_tests[code]
            self.update_selected_tests_display()
            self.filter_inv_catalogue()  # Update highlighting
            self.inv_recalc()
    
    def inv_lookup_patient(self):
        q = self.inv_patient_id.text()
        if not q: return
        p = patient_cms_db.get_patient(q)
        if not p:
            p = patient_cms_db.get_patient_by_phone(q)
        if p:
            self.inv_current_patient = p
            self.inv_patient_info.setText(f"Name: {p['name']}\nID: {p['patientId']}\nPhone: {p['phone']}\nAge/Sex: {p.get('age', 'N/A')}/{p.get('sex', 'N/A')}")
            self.inv_gen_btn.setEnabled(len(self.inv_selected_tests) > 0)
        else:
            QtWidgets.QMessageBox.warning(self, 'Error', 'Patient not found')
    
    def inv_home_toggled(self):
        self.inv_home_fee.setEnabled(self.inv_home_check.isChecked())
    
    def inv_recalc(self):
        subtotal = sum(t['testFees'] for t in self.inv_selected_tests.values())
        home = self.inv_home_fee.value() if self.inv_home_check.isChecked() else 0
        total_before = subtotal + home
        disc_pct = self.inv_discount.value()
        disc_amt = total_before * (disc_pct / 100.0)
        total_after = total_before - disc_amt
        rounded = round(total_after)
        roundoff = rounded - total_after
        
        self.inv_subtotal.setText(f'Subtotal: ₹{subtotal:.2f}')
        self.inv_disc_amt.setText(f'Discount: -₹{disc_amt:.2f}')
        self.inv_roundoff.setText(f'Round Off: ₹{roundoff:.2f}')
        self.inv_total.setText(f'TOTAL: ₹{rounded:.2f}')
    
    def inv_generate(self):
        if not self.inv_current_patient or not self.inv_selected_tests:
            QtWidgets.QMessageBox.warning(self, 'Error', 'Select patient and tests')
            return
        
        try:
            items = list(self.inv_selected_tests.values())
            subtotal = sum(t['testFees'] for t in items)
            home = self.inv_home_fee.value() if self.inv_home_check.isChecked() else 0
            total_before = subtotal + home
            disc_pct = self.inv_discount.value()
            disc_amt = total_before * (disc_pct / 100.0)
            total_after = total_before - disc_amt
            rounded = round(total_after)
            roundoff = rounded - total_after
            
            data = {
                'patient': self.inv_current_patient,
                'items': items,
                'home_collection_fee': home,
                'discount_percentage': disc_pct,
                'is_paid': True,  # Invoices are always created as PAID
                'subtotal': total_before,
                'discount_amount': disc_amt,
                'round_off': roundoff,
                'final_total': rounded
            }
            
            res = invoice_service.create_and_save_invoice(data, self.user.get('id'), CLINIC_ADDRESS, CLINIC_CONTACT, INVOICE_STORAGE_DIR)
            QtWidgets.QMessageBox.information(self, 'Success', f'Invoice {res["invoice_number"]} created')
            webbrowser.open(f"file://{res['filepath']}")
            
            # Clear all invoice form data (but preserve catalogue/special tests)
            self.inv_current_patient = None
            self.inv_patient_id.clear()
            self.inv_patient_info.setText('Not Selected')  # Clear patient details display
            self.inv_selected_tests.clear()
            self.update_selected_tests_display()
            self.inv_home_check.setChecked(False)
            self.inv_home_fee.setValue(0)
            self.inv_discount.setValue(0)
            self.inv_recalc()  # This will update all billing displays
            self.inv_gen_btn.setEnabled(False)
            
            # Refresh datasheet and reports on new invoice
            self.reload_datasheet()
            self.refresh_reports_data()
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, 'Error', str(e))
    
    # ===== PATIENT CMS TAB =====
    def init_cms_tab(self):
        self.init_cms_tab_impl(self.cms_tab)
    
    def init_cms_tab_impl(self, tab_widget):
        """Shared implementation for CMS tab (used in both Pathology and Polyclinic modes)"""
        layout = QtWidgets.QHBoxLayout(tab_widget)
        
        # Left: register form
        left = QtWidgets.QGroupBox('Register Patient')
        left_layout = QtWidgets.QFormLayout(left)
        
        cms_name = QtWidgets.QLineEdit()
        cms_sex = QtWidgets.QComboBox()
        cms_sex.addItems(['M', 'F', 'O'])
        cms_age = QtWidgets.QSpinBox()
        cms_age.setRange(0, 150)
        cms_phone = QtWidgets.QLineEdit()
        cms_email = QtWidgets.QLineEdit()
        cms_address = QtWidgets.QPlainTextEdit()
        
        cms_register_btn = QtWidgets.QPushButton('Register')
        MainWindow.style_button_with_dynamic_spacing(cms_register_btn, font_size=11, padding="8px 16px")
        
        left_layout.addRow('Name:', cms_name)
        left_layout.addRow('Sex:', cms_sex)
        left_layout.addRow('Age:', cms_age)
        left_layout.addRow('Phone:', cms_phone)
        left_layout.addRow('Email:', cms_email)
        left_layout.addRow('Address:', cms_address)
        left_layout.addRow(cms_register_btn)
        
        layout.addWidget(left, 1)
        
        # Right: patient list
        right_w = QtWidgets.QWidget()
        right_layout = QtWidgets.QVBoxLayout(right_w)
        h = QtWidgets.QHBoxLayout()
        h.addWidget(QtWidgets.QLabel('Patients:'))
        cms_search = QtWidgets.QLineEdit()
        cms_search.setPlaceholderText('Search...')
        h.addWidget(cms_search)
        
        # Refresh button for patient CMS
        cms_refresh_btn = QtWidgets.QPushButton('🔄 Refresh')
        cms_refresh_btn.setStyleSheet('font-size: 10px; padding: 4px 8px;')
        cms_refresh_btn.clicked.connect(lambda: cms_reload())
        h.addWidget(cms_refresh_btn)
        
        right_layout.addLayout(h)
        
        cms_table = QtWidgets.QTableWidget()
        cms_table.setColumnCount(7)
        cms_table.setHorizontalHeaderLabels(PATIENT_TABLE_HEADERS)
        right_layout.addWidget(cms_table)
        
        layout.addWidget(right_w, 2)
        
        # Helper functions for this CMS instance
        def cms_register():
            try:
                nid = patient_cms_db.add_patient({
                    'name': cms_name.text(),
                    'sex': cms_sex.currentText(),
                    'age': cms_age.value(),
                    'phone': cms_phone.text(),
                    'email': cms_email.text(),
                    'address': cms_address.toPlainText()
                })
                QtWidgets.QMessageBox.information(self, 'Success', f'Patient {nid} created')
                cms_name.clear(); cms_phone.clear(); cms_email.clear(); cms_address.clear()
                cms_reload()
            except Exception as e:
                QtWidgets.QMessageBox.warning(self, 'Error', str(e))
        
        def cms_reload():
            cms_all_patients = patient_cms_db.get_all_patients()
            cms_filter_patients()
        
        def cms_edit_patient(patient_data):
            """Show dialog to edit patient details"""
            dialog = QtWidgets.QDialog(self)
            dialog.setWindowTitle(f"Edit Patient - {patient_data['name']}")
            layout = QtWidgets.QFormLayout(dialog)
            
            e_name = QtWidgets.QLineEdit(patient_data['name'])
            e_sex = QtWidgets.QComboBox()
            e_sex.addItems(['M', 'F', 'O'])
            e_sex.setCurrentText(patient_data.get('sex', 'M'))
            e_age = QtWidgets.QSpinBox()
            e_age.setRange(0, 150)
            e_age.setValue(patient_data.get('age', 0))
            e_phone = QtWidgets.QLineEdit(patient_data['phone'])
            e_email = QtWidgets.QLineEdit(patient_data.get('email', ''))
            e_address = QtWidgets.QPlainTextEdit(patient_data.get('address', ''))
            
            layout.addRow('Name:', e_name)
            layout.addRow('Sex:', e_sex)
            layout.addRow('Age:', e_age)
            layout.addRow('Phone:', e_phone)
            layout.addRow('Email:', e_email)
            layout.addRow('Address:', e_address)
            
            btn_box = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Save | QtWidgets.QDialogButtonBox.Cancel)
            btn_box.accepted.connect(dialog.accept)
            btn_box.rejected.connect(dialog.reject)
            layout.addRow(btn_box)
            
            if dialog.exec() == QtWidgets.QDialog.Accepted:
                try:
                    new_data = {
                        'name': e_name.text(),
                        'sex': e_sex.currentText(),
                        'age': e_age.value(),
                        'phone': e_phone.text(),
                        'email': e_email.text(),
                        'address': e_address.toPlainText()
                    }
                    patient_cms_db.update_patient(patient_data['patientId'], new_data)
                    QtWidgets.QMessageBox.information(self, 'Success', 'Patient updated successfully')
                    cms_reload()
                except Exception as e:
                    QtWidgets.QMessageBox.warning(self, 'Error', str(e))
        
        def cms_delete_patient(patient_id):
            """Handle patient deletion"""
            reply = QtWidgets.QMessageBox.question(
                self,
                'Confirm Delete',
                f'Are you sure you want to delete patient {patient_id}?\nThis cannot be undone.',
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
                QtWidgets.QMessageBox.No
            )
            
            if reply == QtWidgets.QMessageBox.Yes:
                try:
                    patient_cms_db.delete_patient(patient_id)
                    QtWidgets.QMessageBox.information(self, 'Success', 'Patient deleted successfully')
                    cms_reload()
                except Exception as e:
                    QtWidgets.QMessageBox.critical(self, 'Error', f'Failed to delete patient: {str(e)}')

        def cms_filter_patients():
            q = cms_search.text().lower()
            cms_table.setRowCount(0)
            all_patients = patient_cms_db.get_all_patients()
            for p in all_patients:
                if q in (p.get('name', '') + p.get('patientId', '') + p.get('phone', '')).lower():
                    r = cms_table.rowCount()
                    cms_table.insertRow(r)
                    cms_table.setItem(r, 0, QtWidgets.QTableWidgetItem(p.get('patientId', '')))
                    cms_table.setItem(r, 1, QtWidgets.QTableWidgetItem(p.get('name', '')))
                    cms_table.setItem(r, 2, QtWidgets.QTableWidgetItem(p.get('phone', '')))
                    cms_table.setItem(r, 3, QtWidgets.QTableWidgetItem(str(p.get('age', ''))))
                    cms_table.setItem(r, 4, QtWidgets.QTableWidgetItem(p.get('email', '')))
                    cms_table.setItem(r, 5, QtWidgets.QTableWidgetItem(p.get('address', '')))
                    
                    # Action column
                    action_widget = QtWidgets.QWidget()
                    action_layout = QtWidgets.QHBoxLayout(action_widget)
                    action_layout.setContentsMargins(2, 2, 2, 2)
                    
                    # Edit button (Admin only)
                    if self.user.get('role') == 'admin':
                        edit_btn = QtWidgets.QPushButton('Edit')
                        edit_btn.setStyleSheet('background-color: #0078D4; color: white; padding: 4px;')
                        edit_btn.clicked.connect(lambda _, pat=p: cms_edit_patient(pat))
                        action_layout.addWidget(edit_btn)
                        
                        del_btn = QtWidgets.QPushButton('Delete')
                        del_btn.setStyleSheet('background-color: #D32F2F; color: white; padding: 4px; margin-left: 2px;')
                        del_btn.clicked.connect(lambda _, pid=p.get('patientId'): cms_delete_patient(pid))
                        action_layout.addWidget(del_btn)
                    
                    action_layout.addStretch()
                    cms_table.setCellWidget(r, 6, action_widget)
        
        # Connect signals
        cms_register_btn.clicked.connect(cms_register)
        cms_search.textChanged.connect(cms_filter_patients)
        
        # Initial load
        cms_reload()
    
    
    
    # ===== DATASHEET TAB =====
    def init_datasheet_tab(self):
        layout = QtWidgets.QVBoxLayout(self.datasheet_tab)
        h = QtWidgets.QHBoxLayout()
        h.addWidget(QtWidgets.QLabel('Invoices:'))
        h.addStretch()
        btn_csv = QtWidgets.QPushButton('Export CSV')
        MainWindow.style_button_with_dynamic_spacing(btn_csv, font_size=10, padding="6px 12px")
        btn_csv.clicked.connect(self.export_csv)
        btn_xlsx = QtWidgets.QPushButton('Export XLSX')
        MainWindow.style_button_with_dynamic_spacing(btn_xlsx, font_size=10, padding="6px 12px")
        btn_xlsx.clicked.connect(self.export_xlsx)
        h.addWidget(btn_csv); h.addWidget(btn_xlsx)
        layout.addLayout(h)
        
        self.data_table = QtWidgets.QTableWidget()
        layout.addWidget(self.data_table)
        self.reload_datasheet()
    
    def reload_datasheet(self):
        records = datasheet_db.get_all_invoice_records(full=True)
        is_admin = self.user.get('role') == 'admin'
        
        if not records:
            self.data_table.setRowCount(0)
            return
        
        keys = list(records[0].keys())
        col_count = len(keys) + (1 if is_admin else 0)
        self.data_table.setColumnCount(col_count)
        headers = keys + (['Action'] if is_admin else [])
        self.data_table.setHorizontalHeaderLabels(headers)
        self.data_table.setRowCount(len(records))
        
        for r, rec in enumerate(records):
            for c, k in enumerate(keys):
                self.data_table.setItem(r, c, QtWidgets.QTableWidgetItem(str(rec.get(k, ''))))
            
            # Add delete button for admin only
            if is_admin:
                del_btn = QtWidgets.QPushButton('Delete')
                del_btn.setStyleSheet('font-size: 10px; padding: 2px 6px; color: red;')
                invoice_id = rec.get('invoiceId', '')
                del_btn.clicked.connect(lambda checked, iid=invoice_id: self.delete_datasheet_record(iid))
                self.data_table.setCellWidget(r, len(keys), del_btn)
    
    def delete_datasheet_record(self, invoice_id: str):
        """Admin-only: Delete a record from the datasheet."""
        if self.user.get('role') != 'admin':
            return
        
        reply = QtWidgets.QMessageBox.question(
            self, 'Confirm Delete',
            f'Are you sure you want to delete invoice {invoice_id}?',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No
        )
        if reply == QtWidgets.QMessageBox.Yes:
            datasheet_db.delete_invoice_record(invoice_id)
            self.reload_datasheet()
            QtWidgets.QMessageBox.information(self, 'Deleted', f'Invoice {invoice_id} removed from datasheet.')
    
    def export_csv(self):
        records = datasheet_db.get_all_invoice_records(full=True)
        if not records: return
        default_filename = DEFAULT_CSV_FILENAME.replace('{date}', datetime.date.today().strftime('%Y%m%d'))
        fname, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Save CSV', default_filename, 'CSV (*.csv)')
        if not fname: return
        headers = list(records[0].keys())
        with open(fname, 'w') as f:
            f.write(','.join(headers) + '\n')
            for rec in records:
                f.write(','.join(str(rec.get(h, '')) for h in headers) + '\n')
        QtWidgets.QMessageBox.information(self, 'Success', f'Saved {fname}')
    
    def export_xlsx(self):
        records = datasheet_db.get_all_invoice_records(full=True)
        if not records: return
        default_filename = DEFAULT_XLSX_FILENAME.replace('{date}', datetime.date.today().strftime('%Y%m%d'))
        fname, _ = QtWidgets.QFileDialog.getSaveFileName(self, 'Save XLSX', default_filename, 'Excel (*.xlsx)')
        if not fname: return
        try:
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = list(records[0].keys())
            ws.append(headers)
            for rec in records:
                ws.append([rec.get(h) for h in headers])
            wb.save(fname)
            QtWidgets.QMessageBox.information(self, 'Success', f'Saved {fname}')
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Error', str(e))
    
    # ===== REPORT TRACKER TAB =====
    def init_reports_tab(self):
        # Clear old layout if it exists
        if self.reports_tab.layout():
            while self.reports_tab.layout().count():
                widget = self.reports_tab.layout().takeAt(0).widget()
                if widget:
                    widget.deleteLater()
        
        layout = QtWidgets.QVBoxLayout(self.reports_tab)
        h = QtWidgets.QHBoxLayout()
        h.addWidget(QtWidgets.QLabel('Search:'))
        search = QtWidgets.QLineEdit()
        search.setPlaceholderText('Search...')
        h.addWidget(search)
        layout.addLayout(h)
        
        self.reports_table = QtWidgets.QTableWidget()
        self.reports_table.setColumnCount(7)
        self.reports_table.setHorizontalHeaderLabels(REPORT_TABLE_HEADERS)
        layout.addWidget(self.reports_table)
        
        def filter_reports():
            q = search.text().lower()
            self.reports_table.setRowCount(0)
            for rpt in report_tracker_db.get_all_reports():
                if q in (rpt.get('invoiceId', '') + rpt.get('patientName', '') + rpt.get('patientId', '') + (rpt.get('vid') or '')).lower():
                    r = self.reports_table.rowCount()
                    self.reports_table.insertRow(r)
                    self.reports_table.setItem(r, 0, QtWidgets.QTableWidgetItem(rpt.get('invoiceId', '')))
                    self.reports_table.setItem(r, 1, QtWidgets.QTableWidgetItem(rpt.get('patientName', '')))
                    self.reports_table.setItem(r, 2, QtWidgets.QTableWidgetItem(rpt.get('patientId', '')))
                    self.reports_table.setItem(r, 3, QtWidgets.QTableWidgetItem(rpt.get('status', '')))
                    self.reports_table.setItem(r, 4, QtWidgets.QTableWidgetItem(rpt.get('vid') or ''))
                    self.reports_table.setItem(r, 5, QtWidgets.QTableWidgetItem(rpt.get('created_at', '')))
                    
                    # Create button widget with both Open and Mark Delivered
                    btn_widget = QtWidgets.QWidget()
                    btn_layout = QtWidgets.QHBoxLayout(btn_widget)
                    btn_layout.setContentsMargins(2, 2, 2, 2)
                    
                    open_btn = QtWidgets.QPushButton('Open')
                    open_btn.setMaximumWidth(70)
                    open_btn.clicked.connect(lambda _, fn=rpt.get('pdf_filename'): webbrowser.open(f"file://{os.path.abspath(os.path.join(INVOICE_STORAGE_DIR, fn))}"))
                    btn_layout.addWidget(open_btn)
                    
                    # Mark Delivered button - only if status is not already "DELIVERED"
                    status = rpt.get('status', '').upper()
                    if status != 'DELIVERED':
                        mark_btn = QtWidgets.QPushButton('Mark Delivered')
                        mark_btn.setMaximumWidth(100)
                        mark_btn.setStyleSheet('background-color: #4CAF50; color: white;')
                        mark_btn.clicked.connect(lambda _, inv_id=rpt.get('invoiceId'): self.mark_report_delivered(inv_id))
                        btn_layout.addWidget(mark_btn)
                    
                    # Delete button (Admin only)
                    if self.user.get('role') == 'admin':
                        del_btn = QtWidgets.QPushButton('Delete')
                        del_btn.setMaximumWidth(60)
                        del_btn.setStyleSheet('background-color: #D32F2F; color: white; margin-left: 5px;')
                        del_btn.clicked.connect(lambda _, inv_id=rpt.get('invoiceId'): self.delete_report_ui(inv_id))
                        btn_layout.addWidget(del_btn)
                    
                    btn_layout.addStretch()
                    self.reports_table.setCellWidget(r, 6, btn_widget)
        
        search.textChanged.connect(filter_reports)
        filter_reports()
    
    def delete_report_ui(self, invoice_id: str):
        """Handle report deletion from UI"""
        reply = QtWidgets.QMessageBox.question(
            self,
            'Confirm Delete',
            f'Are you sure you want to delete the report for Invoice #{invoice_id}?\nThis will delete the record and the PDF file.',
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No
        )
        
        if reply == QtWidgets.QMessageBox.Yes:
            try:
                # Delete from DB and get filename
                pdf_filename = report_tracker_db.delete_report(invoice_id)
                
                # Delete file if exists
                if pdf_filename:
                    file_path = os.path.join(INVOICE_STORAGE_DIR, pdf_filename)
                    if os.path.exists(file_path):
                        try:
                            os.remove(file_path)
                            print(f"Deleted file: {file_path}")
                        except Exception as e:
                            print(f"Error deleting file {file_path}: {e}")
                            QtWidgets.QMessageBox.warning(self, "Warning", f"Record deleted but failed to delete file: {e}")
                
                # Refresh table
                self.refresh_reports_data()
                QtWidgets.QMessageBox.information(self, "Success", "Report deleted successfully")
                
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, "Error", f"Failed to delete report: {e}")

    def refresh_reports_data(self):
        """Refresh reports table without reinitializing the entire tab"""
        if hasattr(self, 'reports_table'):
            q = ''
            self.reports_table.setRowCount(0)
            for rpt in report_tracker_db.get_all_reports():
                r = self.reports_table.rowCount()
                self.reports_table.insertRow(r)
                self.reports_table.setItem(r, 0, QtWidgets.QTableWidgetItem(rpt.get('invoiceId', '')))
                self.reports_table.setItem(r, 1, QtWidgets.QTableWidgetItem(rpt.get('patientName', '')))
                self.reports_table.setItem(r, 2, QtWidgets.QTableWidgetItem(rpt.get('patientId', '')))
                self.reports_table.setItem(r, 3, QtWidgets.QTableWidgetItem(rpt.get('status', '')))
                self.reports_table.setItem(r, 4, QtWidgets.QTableWidgetItem(rpt.get('vid') or ''))
                self.reports_table.setItem(r, 5, QtWidgets.QTableWidgetItem(rpt.get('created_at', '')))
                
                # Create button widget with both Open and Mark Delivered
                btn_widget = QtWidgets.QWidget()
                btn_layout = QtWidgets.QHBoxLayout(btn_widget)
                btn_layout.setContentsMargins(2, 2, 2, 2)
                
                open_btn = QtWidgets.QPushButton('Open')
                open_btn.setMaximumWidth(70)
                open_btn.clicked.connect(lambda _, fn=rpt.get('pdf_filename'): webbrowser.open(f"file://{os.path.abspath(os.path.join(INVOICE_STORAGE_DIR, fn))}"))
                btn_layout.addWidget(open_btn)
                
                # Mark Delivered button - only if status is not already "DELIVERED"
                status = rpt.get('status', '').upper()
                if status != 'DELIVERED':
                    mark_btn = QtWidgets.QPushButton('Mark Delivered')
                    mark_btn.setMaximumWidth(100)
                    mark_btn.setStyleSheet('background-color: #4CAF50; color: white;')
                    mark_btn.clicked.connect(lambda _, inv_id=rpt.get('invoiceId'): self.mark_report_delivered(inv_id))
                    btn_layout.addWidget(mark_btn)
                
                # Delete button (Admin only)
                if self.user.get('role') == 'admin':
                    del_btn = QtWidgets.QPushButton('Delete')
                    del_btn.setMaximumWidth(60)
                    del_btn.setStyleSheet('background-color: #D32F2F; color: white; margin-left: 5px;')
                    del_btn.clicked.connect(lambda _, inv_id=rpt.get('invoiceId'): self.delete_report_ui(inv_id))
                    btn_layout.addWidget(del_btn)
                
                btn_layout.addStretch()
                self.reports_table.setCellWidget(r, 6, btn_widget)
    
    def mark_report_delivered(self, invoice_id: str):
        """Show dialog to mark a report as delivered with VID input"""
        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle('Mark Report as Delivered')
        dialog.setGeometry(100, 100, 400, 150)
        
        layout = QtWidgets.QVBoxLayout(dialog)
        
        layout.addWidget(QtWidgets.QLabel(f'Invoice ID: {invoice_id}'))
        
        # VID Input
        vid_layout = QtWidgets.QHBoxLayout()
        vid_layout.addWidget(QtWidgets.QLabel('VID (15 digits):'))
        vid_input = QtWidgets.QLineEdit()
        vid_input.setPlaceholderText('Enter 15-digit VID')
        # Use regex validator to allow up to 15 digits
        vid_regex = QtCore.QRegularExpression(r'[0-9]{0,15}')
        vid_validator = QtGui.QRegularExpressionValidator(vid_regex)
        vid_input.setValidator(vid_validator)
        vid_layout.addWidget(vid_input)
        layout.addLayout(vid_layout)
        
        # Buttons
        btn_layout = QtWidgets.QHBoxLayout()
        mark_btn = QtWidgets.QPushButton('Mark Delivered')
        MainWindow.style_button_with_dynamic_spacing(mark_btn, font_size=11, padding="8px 16px")
        mark_btn.setStyleSheet(mark_btn.styleSheet() + ' background-color: #4CAF50; color: white;')
        cancel_btn = QtWidgets.QPushButton('Cancel')
        MainWindow.style_button_with_dynamic_spacing(cancel_btn, font_size=11, padding="8px 16px")
        
        def do_mark_delivered():
            vid = vid_input.text().strip()
            if not vid:
                QtWidgets.QMessageBox.warning(dialog, 'Input Error', 'Please enter a VID.')
                return
            
            if len(vid) != 15:
                QtWidgets.QMessageBox.warning(dialog, 'Input Error', 'VID must be exactly 15 digits.')
                return
            
            if not vid.isdigit():
                QtWidgets.QMessageBox.warning(dialog, 'Input Error', 'VID must contain only digits.')
                return
            
            try:
                report_tracker_db.mark_report_delivered(invoice_id, vid)
                QtWidgets.QMessageBox.information(dialog, 'Success', f'Report {invoice_id} marked as delivered.')
                dialog.accept()
                # Refresh the reports table and datasheet
                self.refresh_reports_data()
                self.reload_datasheet()
            except Exception as e:
                QtWidgets.QMessageBox.critical(dialog, 'Error', f'Failed to mark report as delivered: {str(e)}')
        
        mark_btn.clicked.connect(do_mark_delivered)
        cancel_btn.clicked.connect(dialog.reject)
        
        btn_layout.addWidget(mark_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)
        
        dialog.exec()
    
    # ===== ADMIN TAB =====
    def init_admin_tab(self):
        layout = QtWidgets.QVBoxLayout(self.admin_tab)
        tabs = QtWidgets.QTabWidget()
        layout.addWidget(tabs)
        
        # Users tab
        users_w = QtWidgets.QWidget()
        users_layout = QtWidgets.QHBoxLayout(users_w)
        
        left = QtWidgets.QGroupBox('Create User')
        left_layout = QtWidgets.QFormLayout(left)
        self.adm_full_name = QtWidgets.QLineEdit()
        self.adm_username = QtWidgets.QLineEdit()
        self.adm_password = QtWidgets.QLineEdit()
        self.adm_password.setEchoMode(QtWidgets.QLineEdit.Password)
        self.adm_role = QtWidgets.QComboBox()
        self.adm_role.addItems(['user', 'admin'])
        btn = QtWidgets.QPushButton('Create')
        MainWindow.style_button_with_dynamic_spacing(btn, font_size=11, padding="8px 16px")
        btn.clicked.connect(self.adm_create_user)
        left_layout.addRow('Full Name:', self.adm_full_name)
        left_layout.addRow('Username:', self.adm_username)
        left_layout.addRow('Password:', self.adm_password)
        left_layout.addRow('Role:', self.adm_role)
        left_layout.addRow(btn)
        users_layout.addWidget(left, 1)
        
        right = QtWidgets.QWidget()
        right_layout = QtWidgets.QVBoxLayout(right)
        right_layout.addWidget(QtWidgets.QLabel('Users:'))
        self.adm_users_table = QtWidgets.QTableWidget()
        self.adm_users_table.setColumnCount(3)
        self.adm_users_table.setHorizontalHeaderLabels(['ID', 'Username', 'Full Name'])
        right_layout.addWidget(self.adm_users_table)
        users_layout.addWidget(right, 2)
        
        tabs.addTab(users_w, 'Users')
        self.adm_reload_users()
        
        # Logs tab
        logs_w = QtWidgets.QWidget()
        logs_layout = QtWidgets.QVBoxLayout(logs_w)
        self.adm_logs_table = QtWidgets.QTableWidget()
        self.adm_logs_table.setColumnCount(3)
        self.adm_logs_table.setHorizontalHeaderLabels(['Timestamp', 'Event', 'User'])
        logs_layout.addWidget(self.adm_logs_table)
        tabs.addTab(logs_w, 'Logs')
        self.adm_reload_logs()
        
        # Backup & Maintenance tab
        maint_w = QtWidgets.QWidget()
        maint_layout = QtWidgets.QVBoxLayout(maint_w)
        maint_group = QtWidgets.QGroupBox("Database Backup")
        mg_layout = QtWidgets.QVBoxLayout(maint_group)
        
        backup_info = QtWidgets.QLabel("Creates a ZIP backup of all system databases. Use this before updating or migrating.")
        backup_info.setWordWrap(True)
        mg_layout.addWidget(backup_info)
        
        self.backup_btn = QtWidgets.QPushButton("Backup Databases Now")
        MainWindow.style_button_with_dynamic_spacing(self.backup_btn, font_size=12, padding="10px 20px")
        self.backup_btn.clicked.connect(self.adm_backup_databases)
        mg_layout.addWidget(self.backup_btn)
        
        # Restore Section
        mg_layout.addSpacing(20)
        restore_label = QtWidgets.QLabel("Restore System")
        restore_label.setStyleSheet("font-weight: bold; margin-top: 10px;")
        mg_layout.addWidget(restore_label)
        
        restore_info = QtWidgets.QLabel("Restores databases, configuration, and logos from a backup ZIP. WARNING: This overwrites current data.")
        restore_info.setStyleSheet("color: #D32F2F;") # Red warning color
        restore_info.setWordWrap(True)
        mg_layout.addWidget(restore_info)
        
        self.restore_btn = QtWidgets.QPushButton("Restore from Backup")
        MainWindow.style_button_with_dynamic_spacing(self.restore_btn, font_size=12, padding="10px 20px")
        self.restore_btn.setStyleSheet("""
            QPushButton { 
                background-color: #f44336; 
                color: white; 
                font-weight: bold; 
                border-radius: 4px;
            } 
            QPushButton:hover { background-color: #d32f2f; }
        """)
        self.restore_btn.clicked.connect(self.adm_restore_backup)
        mg_layout.addWidget(self.restore_btn)
        
        maint_layout.addWidget(maint_group)
        maint_layout.addStretch()
        
        tabs.addTab(maint_w, 'Maintenance')
        
    def adm_restore_backup(self):
        """Restore system from backup via Admin Panel"""
        zip_path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Select Backup Archive", "", "ZIP Files (*.zip)")
        if zip_path:
            from app.restore_utils import restore_system_from_backup
            success = restore_system_from_backup(zip_path, self)
            if success:
                # Logic to restart or close?
                # The helper shows a message saying app will close.
                QtWidgets.QApplication.quit()
                sys.exit(0)
    
    def adm_create_user(self):
        try:
            auth_db.create_user(self.adm_username.text(), self.adm_password.text(), self.adm_full_name.text(), self.adm_role.currentText())
            QtWidgets.QMessageBox.information(self, 'Success', 'User created')
            self.adm_username.clear(); self.adm_password.clear(); self.adm_full_name.clear()
            self.adm_reload_users()
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, 'Error', str(e))
    
    def adm_backup_databases(self):
        """Create a zip backup of all databases"""
        try:
            from app.utils import get_database_dir
            import shutil
            from datetime import datetime
            
            db_dir = get_database_dir()
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_name = f"manual_backup_{timestamp}"
            backup_path = os.path.join(db_dir, "backups", backup_name)
            
            # Create backup directory
            os.makedirs(backup_path, exist_ok=True)
            
            # Copy all .db files
            count = 0
            for file in os.listdir(db_dir):
                if file.endswith(".db"):
                    shutil.copy2(os.path.join(db_dir, file), os.path.join(backup_path, file))
                    count += 1
            
            # --- Backup Config & Logos ---
            project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) # app -> root
            if getattr(sys, 'frozen', False):
                 project_root = os.path.dirname(sys.executable)
                 
            conf_path = get_config_path()
            if os.path.exists(conf_path):
                shutil.copy2(conf_path, os.path.join(backup_path, "config.yaml"))
                count += 1
                
                # Check for custom logos in config
                try:
                    with open(conf_path, 'r', encoding='utf-8') as f:
                        config = yaml.safe_load(f) or {}
                        
                    logos_to_backup = []
                    if 'BRANDING_LOGO' in config and config['BRANDING_LOGO']:
                        logos_to_backup.append(config['BRANDING_LOGO'])
                    if 'REPORT_LOGO' in config and config['REPORT_LOGO']:
                        logos_to_backup.append(config['REPORT_LOGO'])
                        
                    if logos_to_backup:
                        logo_backup_dir = os.path.join(backup_path, "logos")
                        os.makedirs(logo_backup_dir, exist_ok=True)
                        
                        for logo_file in logos_to_backup:
                            # Try to resolve path (could be absolute or relative to assets)
                            src_path = None
                            if os.path.isabs(logo_file):
                                if os.path.exists(logo_file):
                                    src_path = logo_file
                            else:
                                # Try assets dir
                                candidate = get_asset_path(logo_file)
                                if os.path.exists(candidate):
                                    src_path = candidate
                                    
                            if src_path:
                                shutil.copy2(src_path, os.path.join(logo_backup_dir, os.path.basename(logo_file)))
                                count += 1
                except Exception as ex:
                    print(f"Warning backing up logos: {ex}")
            
            # Create ZIP archive
            zip_path = shutil.make_archive(backup_path, 'zip', backup_path)
            
            # Clean up raw folder to save space, keeping only the zip
            shutil.rmtree(backup_path)
            
            msg = f"Backup created successfully!\n\nFiles backed up: {count}\nLocation: {zip_path}"
            
            # Open folder in explorer
            if os.name == 'nt':
                os.startfile(os.path.dirname(zip_path))
            
            QtWidgets.QMessageBox.information(self, "Backup Success", msg)
            
            # Log action if auth_db is available
            try:
                auth_db.log_event(self.user['username'], "backup_created", f"Created backup: {os.path.basename(zip_path)}")
                self.adm_reload_logs()
            except:
                pass
                
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Backup Failed", f"An error occurred during backup:\n{str(e)}")

    def adm_reload_users(self):
        users = auth_db.get_all_users()
        self.adm_users_table.setRowCount(len(users))
        for r, u in enumerate(users):
            self.adm_users_table.setItem(r, 0, QtWidgets.QTableWidgetItem(str(u['id'])))
            self.adm_users_table.setItem(r, 1, QtWidgets.QTableWidgetItem(u['username']))
            self.adm_users_table.setItem(r, 2, QtWidgets.QTableWidgetItem(u.get('full_name', '')))
    
    def adm_reload_logs(self):
        logs = auth_db.get_login_logs()
        self.adm_logs_table.setRowCount(len(logs))
        for r, log in enumerate(logs):
            self.adm_logs_table.setItem(r, 0, QtWidgets.QTableWidgetItem(log.get('timestamp', '')))
            self.adm_logs_table.setItem(r, 1, QtWidgets.QTableWidgetItem(log.get('event', '')))
            self.adm_logs_table.setItem(r, 2, QtWidgets.QTableWidgetItem(str(log.get('user_id', ''))))
    
    # ===== POLYCLINIC TAB INITIALIZATION =====
    
    def init_poly_booking_tab(self):
        """Initialize Patient Booking tab for Polyclinic"""
        # Main layout (no scroll area)
        layout = QtWidgets.QVBoxLayout(self.poly_booking_tab)
        layout.setSpacing(10)
        
        # ===== Patient Lookup Section =====
        patient_group = QtWidgets.QGroupBox("Patient Details")
        patient_layout = QtWidgets.QHBoxLayout(patient_group)
        
        patient_layout.addWidget(QtWidgets.QLabel("Patient Phone:"))
        self.poly_patient_phone = QtWidgets.QLineEdit()
        patient_layout.addWidget(self.poly_patient_phone)
        
        search_btn = QtWidgets.QPushButton("Search")
        self.style_button_with_dynamic_spacing(search_btn)
        search_btn.clicked.connect(self.poly_lookup_patient)
        patient_layout.addWidget(search_btn)
        
        search_btn.clicked.connect(self.poly_lookup_patient)
        patient_layout.addWidget(search_btn)
        
        layout.addWidget(patient_group)
        
        # Patient info display
        info_group = QtWidgets.QGroupBox("Patient Information")
        info_layout = QtWidgets.QFormLayout(info_group)
        
        self.poly_patient_name = QtWidgets.QLineEdit()
        self.poly_patient_name.setReadOnly(True)
        info_layout.addRow("Name:", self.poly_patient_name)
        
        self.poly_patient_id = QtWidgets.QLineEdit()
        self.poly_patient_id.setReadOnly(True)
        info_layout.addRow(f"{PATIENT_ID_LABEL}:", self.poly_patient_id)
        
        self.poly_patient_age = QtWidgets.QLineEdit()
        self.poly_patient_age.setReadOnly(True)
        info_layout.addRow("Age:", self.poly_patient_age)
        
        self.poly_patient_gender = QtWidgets.QLineEdit()
        self.poly_patient_gender.setReadOnly(True)
        info_layout.addRow("Gender:", self.poly_patient_gender)
        
        layout.addWidget(info_group)
        
        # ===== Doctor Selection Section =====
        doctor_group = QtWidgets.QGroupBox("Doctor Selection")
        doctor_layout = QtWidgets.QVBoxLayout(doctor_group)
        
        # Search options
        search_layout = QtWidgets.QHBoxLayout()
        search_layout.addWidget(QtWidgets.QLabel("Find Doctor:"))
        
        self.poly_doctor_search = QtWidgets.QLineEdit()
        self.poly_doctor_search.setPlaceholderText("Search by name or speciality...")
        # Don't connect here - will load on first focus
        search_layout.addWidget(self.poly_doctor_search)
        
        # Refresh button for doctors in booking
        booking_refresh_btn = QtWidgets.QPushButton('🔄 Refresh')
        booking_refresh_btn.setStyleSheet('font-size: 10px; padding: 4px 8px;')
        booking_refresh_btn.clicked.connect(self.poly_refresh_booking_doctors)
        search_layout.addWidget(booking_refresh_btn)
        
        doctor_layout.addLayout(search_layout)
        
        # Doctor list (using QListWidget instead of table)
        self.poly_doctor_list = QtWidgets.QListWidget()
        self.poly_doctor_list.itemSelectionChanged.connect(self.poly_on_doctor_selected)
        doctor_layout.addWidget(self.poly_doctor_list)
        
        # Load doctors on first time the search field gets focus
        self.poly_doctor_search.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.poly_doctor_search.focusIn = lambda: self._poly_load_doctors_once()
        self.poly_doctors_loaded = False
        
        # Load doctors on first time the search field gets focus
        self.poly_doctor_search.setFocusPolicy(QtCore.Qt.StrongFocus)
        self.poly_doctor_search.focusIn = lambda: self._poly_load_doctors_once()
        self.poly_doctors_loaded = False
        
        # Limit height of doctor list
        self.poly_doctor_list.setMaximumHeight(120)
        
        layout.addWidget(doctor_group)
        
        # ===== Date & Time Selection (Side-by-Side) =====
        datetime_group = QtWidgets.QGroupBox("Appointment Details")
        datetime_layout = QtWidgets.QHBoxLayout(datetime_group)
        
        # Left: Calendar
        self.poly_booking_date = QtWidgets.QCalendarWidget()
        self.poly_booking_date.setMinimumDate(QtCore.QDate.currentDate())
        self.poly_booking_date.clicked.connect(self.poly_update_time_slots)
        datetime_layout.addWidget(self.poly_booking_date, 1)
        
        # Right: Time, Fees, Serial
        right_panel = QtWidgets.QWidget()
        right_layout = QtWidgets.QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        
        # Time slot
        right_layout.addWidget(QtWidgets.QLabel("Select Time:"))
        self.poly_time_slot = QtWidgets.QListWidget()
        self.poly_time_slot.itemSelectionChanged.connect(self.poly_update_serial_number)
        right_layout.addWidget(self.poly_time_slot)
        
        # Fees
        fees_layout = QtWidgets.QHBoxLayout()
        fees_layout.addWidget(QtWidgets.QLabel("Fees:"))
        self.poly_fees_display = QtWidgets.QLineEdit()
        self.poly_fees_display.setReadOnly(True)
        fees_layout.addWidget(self.poly_fees_display)
        right_layout.addLayout(fees_layout)
        
        # Serial
        serial_layout = QtWidgets.QHBoxLayout()
        serial_layout.addWidget(QtWidgets.QLabel("Serial:"))
        self.poly_serial_display = QtWidgets.QLineEdit()
        self.poly_serial_display.setReadOnly(True)
        serial_layout.addWidget(self.poly_serial_display)
        right_layout.addLayout(serial_layout)
        
        datetime_layout.addWidget(right_panel, 1)
        
        layout.addWidget(datetime_group, 1)
        
        # ===== Book Button =====
        # ===== Book Button =====
        book_btn = QtWidgets.QPushButton("Book Appointment")
        MainWindow.style_button_with_dynamic_spacing(book_btn, font_size=13, padding="8px 16px")
        book_btn.clicked.connect(self.poly_book_appointment)
        layout.addWidget(book_btn)
    
    def poly_lookup_patient(self):
        """Lookup patient by phone number"""
        phone = self.poly_patient_phone.text().strip()
        if not phone:
            QtWidgets.QMessageBox.warning(self, "Error", "Please enter a phone number")
            return
        
        try:
            patient = patient_cms_db.get_patient_by_phone(phone)
            if patient:
                self.poly_patient_name.setText(patient.get('name', ''))
                self.poly_patient_id.setText(str(patient.get('patientId', '')))
                self.poly_patient_age.setText(str(patient.get('age', '')))
                self.poly_patient_gender.setText(patient.get('sex', ''))
                # Store the patientId for booking (this is the primary key)
                self.poly_selected_patient_id = patient.get('patientId')
            else:
                QtWidgets.QMessageBox.information(self, "Not Found", "Patient not found. Please register in Patient CMS first.")
                self.poly_clear_patient_info()
                self.poly_selected_patient_id = None
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Error", f"Error looking up patient: {str(e)}")
    
    def poly_clear_patient_info(self):
        """Clear patient info fields"""
        self.poly_patient_name.clear()
        self.poly_patient_id.clear()
        self.poly_patient_age.clear()
        self.poly_patient_gender.clear()
        self.poly_selected_patient_id = None
    
    def poly_filter_doctors(self):
        """Filter doctors based on search text"""
        # Load doctors once if not already loaded
        self._poly_load_doctors_once()
        
        search_text = self.poly_doctor_search.text().lower()
        try:
            doctors = polyclinic_db.get_all_doctors()
            filtered = [d for d in doctors if search_text in d['name'].lower() or search_text in d['speciality'].lower()]
            
            self.poly_doctor_list.clear()
            for doc in filtered:
                display_text = f"{doc['name']} - {doc['speciality']} ({doc.get('degree', '')}) - ₹{doc.get('visiting_fees', 0)}"
                item = QtWidgets.QListWidgetItem(display_text)
                item.setData(QtCore.Qt.UserRole, doc)  # Store doctor data
                self.poly_doctor_list.addItem(item)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Error", f"Error loading doctors: {str(e)}")
    
    def _poly_load_doctors_once(self):
        """Load doctors list once on first focus"""
        if not self.poly_doctors_loaded:
            self.poly_doctors_loaded = True
            self.poly_doctor_search.textChanged.connect(self.poly_filter_doctors)
            self.poly_filter_doctors()
    
    def poly_on_doctor_selected(self):
        """Handle doctor selection"""
        selected_items = self.poly_doctor_list.selectedItems()
        if not selected_items:
            return
        
        item = selected_items[0]
        doctor = item.data(QtCore.Qt.UserRole)
        
        self.poly_selected_doctor_id = doctor['doctor_id']
        self.poly_selected_doctor = doctor['name']
        self.poly_fees_display.setText(str(doctor.get('visiting_fees', 0)))
        
        # Update calendar to show available days
        self.poly_update_calendar_availability()
        
        # Update time slots for current date
        self.poly_update_time_slots()
    
    def poly_update_calendar_availability(self):
        """Grey out dates where doctor is not available"""
        if not hasattr(self, 'poly_selected_doctor_id'):
            return
            
        try:
            # Get available weekdays (0=Mon, 6=Sun)
            availability = polyclinic_db.get_doctor_availability(self.poly_selected_doctor_id)
            available_weekdays = {a['day_of_week'] for a in availability}
            
            # Format for unavailable days (Greyed out)
            unavailable_fmt = QtGui.QTextCharFormat()
            unavailable_fmt.setForeground(QtGui.QBrush(QtGui.QColor("lightgray")))
            # unavailable_fmt.setBackground(QtGui.QBrush(QtGui.QColor("#f0f0f0")))
            
            # Format for available days (Standard)
            available_fmt = QtGui.QTextCharFormat()
            available_fmt.setForeground(QtGui.QBrush(QtCore.Qt.black))
            available_fmt.setFontWeight(QtGui.QFont.Bold)
            
            # Reset formats first (clear all custom formatting)
            self.poly_booking_date.setDateTextFormat(QtCore.QDate(), QtGui.QTextCharFormat())
            
            # Apply over a 3-month range from current month
            start_date = self.poly_booking_date.selectedDate().addDays(-15) # Start a bit before
            for i in range(90): # Look ahead 90 days
                date = start_date.addDays(i)
                # Python weekday() is 0-6 (Mon-Sun)
                # QDate.dayOfWeek() is 1-7 (Mon-Sun)
                # We use python conversion for consistency with DB
                py_weekday = date.toPython().weekday()
                
                # Check if date is in the past
                is_past = date < QtCore.QDate.currentDate()
                
                if is_past or py_weekday not in available_weekdays:
                    self.poly_booking_date.setDateTextFormat(date, unavailable_fmt)
                else:
                    self.poly_booking_date.setDateTextFormat(date, available_fmt)
                    
        except Exception as e:
            print(f"Error updating calendar: {e}")
    
    def poly_update_time_slots(self):
        """Update available time slots for selected doctor and date"""
        if not hasattr(self, 'poly_selected_doctor_id'):
            QtWidgets.QMessageBox.warning(self, "Error", "Please select a doctor first")
            return
        
        try:
            selected_date_obj = self.poly_booking_date.selectedDate().toPython()
            day_of_week = selected_date_obj.weekday()
            
            # Get availability for this doctor and day
            availability = polyclinic_db.get_doctor_availability(self.poly_selected_doctor_id)
            day_slots = [a for a in availability if a['day_of_week'] == day_of_week]
            
            self.poly_time_slot.clear()
            self.poly_available_slots = []  # Store slot objects for later use
            
            for slot in day_slots:
                display_text = f"{slot['start_time']} - {slot['end_time']}"
                item = QtWidgets.QListWidgetItem(display_text)
                item.setData(QtCore.Qt.UserRole, slot)  # Store slot data
                self.poly_time_slot.addItem(item)
                self.poly_available_slots.append(slot)
            
            if not day_slots:
                self.poly_serial_display.clear()
            else:
                # Update serial number
                self.poly_update_serial_number()
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Error", f"Error updating time slots: {str(e)}")
    
    def poly_update_serial_number(self):
        """Update serial number display for selected doctor, date, and time"""
        if not hasattr(self, 'poly_selected_doctor_id') or self.poly_time_slot.count() == 0 or not self.poly_time_slot.selectedItems():
            self.poly_serial_display.clear()
            return
        
        try:
            selected_item = self.poly_time_slot.selectedItems()[0]
            slot = selected_item.data(QtCore.Qt.UserRole)
            
            selected_date = self.poly_booking_date.selectedDate().toPython()
            booking_date_str = str(selected_date)
            booking_time = slot['start_time']
            
            # Calculate serial number for this doctor-date-time combination
            bookings = polyclinic_db.get_bookings_for_doctor_date_time(self.poly_selected_doctor_id, booking_date_str, booking_time)
            serial = len(bookings) + 1
            
            self.poly_serial_display.setText(str(serial))
        except Exception as e:
            self.poly_serial_display.clear()
    
    def poly_book_appointment(self):
        """Book an appointment"""
        if not hasattr(self, 'poly_selected_patient_id') or not self.poly_selected_patient_id:
            QtWidgets.QMessageBox.warning(self, "Error", "Please search and select a patient first")
            return
        
        if not hasattr(self, 'poly_selected_doctor_id'):
            QtWidgets.QMessageBox.warning(self, "Error", "Please select a doctor")
            return
        
        if not self.poly_time_slot.selectedItems():
            QtWidgets.QMessageBox.warning(self, "Error", "Please select a time slot")
            return
        
        try:
            selected_item = self.poly_time_slot.selectedItems()[0]
            slot = selected_item.data(QtCore.Qt.UserRole)
            
            booking_date = str(self.poly_booking_date.selectedDate().toPython())
            booking_time = slot['start_time']
            serial_number = int(self.poly_serial_display.text())
            
            # Create booking using the stored patient ID
            booking_id = polyclinic_db.add_booking(
                self.poly_selected_patient_id,
                doctor_id=self.poly_selected_doctor_id,
                booking_date=booking_date,
                booking_time=booking_time,
                serial_number=serial_number,
                payment_status='PAID',
                attendance_status='PENDING'
            )
            
            QtWidgets.QMessageBox.information(
                self, 
                "Success", 
                f"Appointment booked successfully!\nSerial Number: {serial_number}\nBooking Date: {booking_date}\nTime: {booking_time}"
            )
            
            # Clear form
            self.poly_patient_phone.clear()
            self.poly_clear_patient_info()
            self.poly_doctor_search.clear()
            self.poly_doctor_list.clear()
            self.poly_time_slot.clear()
            self.poly_fees_display.clear()
            self.poly_serial_display.clear()
            
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Error", f"Error booking appointment: {str(e)}")
    
    def init_poly_doctor_tab(self):
        """Initialize Doctor Entry/Lookup tab"""
        layout = QtWidgets.QVBoxLayout(self.poly_doctor_tab)
        
        is_admin = self.user.get('role') == 'admin'
        
        # ===== Right Panel: Doctor List (Always created) =====
        list_widget = QtWidgets.QWidget()
        list_layout = QtWidgets.QVBoxLayout(list_widget)
        
        search_layout = QtWidgets.QHBoxLayout()
        search_layout.addWidget(QtWidgets.QLabel("Search:"))
        self.poly_doc_search = QtWidgets.QLineEdit()
        search_layout.addWidget(self.poly_doc_search)
        
        # Refresh button for doctor list
        doc_refresh_btn = QtWidgets.QPushButton('🔄 Refresh')
        MainWindow.style_button_with_dynamic_spacing(doc_refresh_btn, font_size=10, padding="4px 8px")
        doc_refresh_btn.clicked.connect(self.poly_refresh_doctor_list)
        search_layout.addWidget(doc_refresh_btn)
        
        list_layout.addLayout(search_layout)
        
        self.poly_doctor_mgmt_table = QtWidgets.QTableWidget()
        cols = ['ID', 'Name', 'Speciality', 'Degree', 'Fees']
        if is_admin:
            cols.extend(['Edit', 'Delete'])
        
        self.poly_doctor_mgmt_table.setColumnCount(len(cols))
        self.poly_doctor_mgmt_table.setHorizontalHeaderLabels(cols)
        self.poly_doctor_mgmt_table.itemSelectionChanged.connect(self.poly_on_doctor_table_selected)
        list_layout.addWidget(self.poly_doctor_mgmt_table)
        
        # Initial load of doctors
        self.poly_reload_doctor_list()
        
        # Connect search
        self.poly_doc_search.textChanged.connect(lambda: self.poly_reload_doctor_list())
        
        if is_admin:
            splitter = QtWidgets.QSplitter(QtCore.Qt.Orientation.Horizontal)
            
            # ===== Left Panel: Doctor Entry Form =====
            form_widget = QtWidgets.QWidget()
            form_layout = QtWidgets.QVBoxLayout(form_widget)
            
            # Basic doctor info
            form_group = QtWidgets.QGroupBox("Add/Edit Doctor")
            form_group_layout = QtWidgets.QFormLayout(form_group)
            
            self.poly_doc_name = QtWidgets.QLineEdit()
            form_group_layout.addRow("Name:", self.poly_doc_name)
            
            self.poly_doc_speciality = QtWidgets.QLineEdit()
            form_group_layout.addRow("Speciality:", self.poly_doc_speciality)
            
            self.poly_doc_degree = QtWidgets.QLineEdit()
            form_group_layout.addRow("Degree:", self.poly_doc_degree)
            
            self.poly_doc_fees = QtWidgets.QDoubleSpinBox()
            self.poly_doc_fees.setRange(0, 10000)
            form_group_layout.addRow("Visiting Fees:", self.poly_doc_fees)
            
            form_layout.addWidget(form_group)
            
            # ===== Per-Day Schedule Section =====
            schedule_group = QtWidgets.QGroupBox("Schedule (Per Day)")
            schedule_layout = QtWidgets.QVBoxLayout(schedule_group)
            
            # Scrollable area for days
            scroll = QtWidgets.QScrollArea()
            scroll.setWidgetResizable(True)
            days_widget = QtWidgets.QWidget()
            self.poly_days_layout = QtWidgets.QVBoxLayout(days_widget)
            
            days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            self.poly_day_widgets = {}
            
            for day_idx, day_name in enumerate(days):
                # Day group box
                day_group = QtWidgets.QGroupBox(day_name)
                day_group_layout = QtWidgets.QVBoxLayout(day_group)
                
                # Checkbox to enable/disable this day
                day_checkbox = QtWidgets.QCheckBox(f"Available on {day_name}")
                day_group_layout.addWidget(day_checkbox)
                
                # Time slots container for this day
                time_slots_container = QtWidgets.QVBoxLayout()
                
                # Add button for time slots
                add_slot_btn = QtWidgets.QPushButton(f"Add Time Slot for {day_name}")
                MainWindow.style_button_with_dynamic_spacing(add_slot_btn, font_size=10, padding="4px 8px")
                add_slot_btn.setEnabled(False)  # Disabled until day is checked
                
                day_group_layout.addWidget(add_slot_btn)
                day_group_layout.addLayout(time_slots_container)
                
                # Store for later reference
                self.poly_day_widgets[day_idx] = {
                    'checkbox': day_checkbox,
                    'group': day_group,
                    'add_btn': add_slot_btn,
                    'slots_layout': time_slots_container,
                    'slots': []
                }
                
                # Connect signals
                day_checkbox.toggled.connect(lambda checked, idx=day_idx: self._poly_on_day_toggled(idx, checked))
                add_slot_btn.clicked.connect(lambda checked=False, idx=day_idx: self._poly_on_add_slot(idx))
                
                self.poly_days_layout.addWidget(day_group)
            
            self.poly_days_layout.addStretch()
            scroll.setWidget(days_widget)
            schedule_layout.addWidget(scroll)
            
            form_layout.addWidget(schedule_group, 1)
            
            # Save button
            save_btn = QtWidgets.QPushButton("Save Doctor")
            MainWindow.style_button_with_dynamic_spacing(save_btn, font_size=12, padding="8px 16px")
            save_btn.clicked.connect(self.poly_save_doctor)
            form_layout.addWidget(save_btn)
            
            splitter.addWidget(form_widget)
            splitter.addWidget(list_widget)
            layout.addWidget(splitter)
        else:
            layout.addWidget(list_widget)
    
    def _poly_on_day_toggled(self, day_idx, is_checked):
        """Handle day checkbox toggle"""
        day_widget = self.poly_day_widgets[day_idx]
        day_widget['add_btn'].setEnabled(is_checked)
        
        if is_checked and len(day_widget['slots']) == 0:
            # Add default slot when day is first enabled
            self._poly_add_slot_impl(day_idx, "09:00", "17:00")
        elif not is_checked:
            # Clear all slots when day is unchecked
            self._poly_clear_day_slots(day_idx)
    
    def _poly_on_add_slot(self, day_idx):
        """Handle add time slot button click"""
        self._poly_add_slot_impl(day_idx, "09:00", "17:00")
    
    def _poly_add_slot_impl(self, day_idx, start_time="09:00", end_time="17:00"):
        """Implementation to add a time slot"""
        day_widget = self.poly_day_widgets[day_idx]
        slot_index = len(day_widget['slots'])
        
        slot_layout = QtWidgets.QHBoxLayout()
        
        # Start time
        start_widget = QtWidgets.QTimeEdit()
        start_widget.setTime(QtCore.QTime.fromString(start_time, "HH:mm"))
        slot_layout.addWidget(QtWidgets.QLabel("From:"))
        slot_layout.addWidget(start_widget)
        
        # End time
        end_widget = QtWidgets.QTimeEdit()
        end_widget.setTime(QtCore.QTime.fromString(end_time, "HH:mm"))
        slot_layout.addWidget(QtWidgets.QLabel("To:"))
        slot_layout.addWidget(end_widget)
        
        # Remove button
        remove_btn = QtWidgets.QPushButton("Remove")
        remove_btn.setMaximumWidth(80)
        remove_btn.setStyleSheet("background-color: #D32F2F; color: white; padding: 2px;")
        remove_btn.clicked.connect(lambda: self._poly_remove_slot_impl(day_idx, slot_index))
        slot_layout.addWidget(remove_btn)
        
        # Store slot
        day_widget['slots'].append({
            'start': start_widget,
            'end': end_widget,
            'layout': slot_layout
        })
        
        day_widget['slots_layout'].addLayout(slot_layout)
    
    def _poly_remove_slot_impl(self, day_idx, slot_index):
        """Implementation to remove a time slot"""
        day_widget = self.poly_day_widgets[day_idx]
        if slot_index < len(day_widget['slots']):
            slot = day_widget['slots'].pop(slot_index)
            layout = slot['layout']
            
            # Remove all widgets from layout
            while layout.count() > 0:
                item = layout.takeAt(0)
                if item and item.widget():
                    item.widget().deleteLater()
            
            # Remove layout from parent
            slots_layout = day_widget['slots_layout']
            for i in range(slots_layout.count()):
                if slots_layout.itemAt(i) == layout:
                    slots_layout.takeAt(i)
                    break
    
    def _poly_clear_day_slots(self, day_idx):
        """Clear all time slots for a day"""
        day_widget = self.poly_day_widgets[day_idx]
        day_widget['slots'].clear()
        while day_widget['slots_layout'].count() > 0:
            item = day_widget['slots_layout'].takeAt(0)
            if item:
                item.deleteLater()
    
    def poly_save_doctor(self):
        """Save doctor and per-day availability"""
        name = self.poly_doc_name.text().strip()
        speciality = self.poly_doc_speciality.text().strip()
        degree = self.poly_doc_degree.text().strip()
        fees = self.poly_doc_fees.value()
        
        if not name or not speciality or not degree:
            QtWidgets.QMessageBox.warning(self, "Error", "Please fill in all doctor details")
            return
        
        # Collect all day schedules
        schedule_data = {}  # day_idx -> [(start_time, end_time), ...]
        for day_idx in range(7):
            day_widget = self.poly_day_widgets[day_idx]
            if day_widget['checkbox'].isChecked():
                if len(day_widget['slots']) > 0:
                    time_slots = []
                    for slot in day_widget['slots']:
                        start_time = slot['start'].time().toString("HH:mm")
                        end_time = slot['end'].time().toString("HH:mm")
                        time_slots.append((start_time, end_time))
                    schedule_data[day_idx] = time_slots
        
        if not schedule_data:
            QtWidgets.QMessageBox.warning(self, "Error", "Please select at least one day with time slots")
            return
        
        try:
            # Check if we're editing an existing doctor
            existing_doctor = None
            all_doctors = polyclinic_db.get_all_doctors()
            for doc in all_doctors:
                if doc['name'].lower() == name.lower() and doc['speciality'].lower() == speciality.lower():
                    existing_doctor = doc
                    break
            
            if existing_doctor:
                # Update existing doctor
                doctor_id = existing_doctor['doctor_id']
                polyclinic_db.update_doctor(doctor_id, {
                    'name': name,
                    'speciality': speciality,
                    'degree': degree,
                    'visiting_fees': fees
                })
                
                # Clear and re-add availability per day
                polyclinic_db.clear_doctor_availability(doctor_id)
                for day_idx, time_slots in schedule_data.items():
                    for start_time, end_time in time_slots:
                        polyclinic_db.add_availability(doctor_id, day_idx, start_time, end_time)
                
                QtWidgets.QMessageBox.information(self, "Success", f"Doctor '{name}' updated successfully!")
            else:
                # Add new doctor
                doctor_id = polyclinic_db.add_doctor(name, speciality, degree, fees)
                
                # Add availability per day
                for day_idx, time_slots in schedule_data.items():
                    for start_time, end_time in time_slots:
                        polyclinic_db.add_availability(doctor_id, day_idx, start_time, end_time)
                
                QtWidgets.QMessageBox.information(self, "Success", f"Doctor '{name}' added successfully!")
            
            # Clear form
            self.poly_doc_name.clear()
            self.poly_doc_speciality.clear()
            self.poly_doc_degree.clear()
            self.poly_doc_fees.setValue(0)
            
            # Clear all day schedules
            for day_idx in range(7):
                day_widget = self.poly_day_widgets[day_idx]
                day_widget['checkbox'].setChecked(False)
                day_widget['slots'].clear()
                while day_widget['slots_layout'].count() > 0:
                    item = day_widget['slots_layout'].takeAt(0)
                    if item:
                        item.deleteLater()
            
            self.poly_reload_doctor_list()
            
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Error", f"Error saving doctor: {str(e)}")
    
    def poly_reload_doctor_list(self):
        """Reload doctor list table"""
        search_text = self.poly_doc_search.text().lower()
        is_admin = self.user.get('role') == 'admin'
        
        try:
            doctors = polyclinic_db.get_all_doctors()
            filtered = [d for d in doctors if search_text in d['name'].lower() or search_text in d['speciality'].lower()]
            
            self.poly_doctor_mgmt_table.setRowCount(len(filtered))
            for r, doc in enumerate(filtered):
                self.poly_doctor_mgmt_table.setItem(r, 0, QtWidgets.QTableWidgetItem(str(doc['doctor_id'])))
                self.poly_doctor_mgmt_table.setItem(r, 1, QtWidgets.QTableWidgetItem(doc['name']))
                self.poly_doctor_mgmt_table.setItem(r, 2, QtWidgets.QTableWidgetItem(doc['speciality']))
                self.poly_doctor_mgmt_table.setItem(r, 3, QtWidgets.QTableWidgetItem(doc.get('degree', '')))
                self.poly_doctor_mgmt_table.setItem(r, 4, QtWidgets.QTableWidgetItem(str(doc.get('visiting_fees', 0))))
                
                if is_admin:
                    # Edit button
                    edit_btn = QtWidgets.QPushButton("Edit")
                    MainWindow.style_button_with_dynamic_spacing(edit_btn, font_size=10, padding="4px 8px")
                    edit_btn.clicked.connect(lambda checked, doc_id=doc['doctor_id']: self.poly_edit_doctor(doc_id))
                    self.poly_doctor_mgmt_table.setCellWidget(r, 5, edit_btn)
                    
                    # Delete button
                    delete_btn = QtWidgets.QPushButton("Delete")
                    delete_btn.setStyleSheet("background-color: #D32F2F; color: white; padding: 4px;")
                    delete_btn.clicked.connect(lambda checked, doc_id=doc['doctor_id']: self.poly_delete_doctor(doc_id))
                    self.poly_doctor_mgmt_table.setCellWidget(r, 6, delete_btn)
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Error", f"Error loading doctors: {str(e)}")
    
    def _poly_load_doctor_list_once(self):
        """Load doctor list once on first focus"""
        if not self.poly_doc_list_loaded:
            self.poly_doc_list_loaded = True
            self.poly_reload_doctor_list()
    
    def poly_on_doctor_table_selected(self):
        """Handle doctor table selection"""
        selected_rows = self.poly_doctor_mgmt_table.selectedIndexes()
        if not selected_rows:
            return
        
        row = selected_rows[0].row()
        doctor_id = int(self.poly_doctor_mgmt_table.item(row, 0).text())
        self.poly_edit_doctor(doctor_id)
    
    def poly_edit_doctor(self, doctor_id):
        """Load doctor into edit form"""
        try:
            doctor = polyclinic_db.get_doctor(doctor_id)
            if not doctor:
                return
            
            self.poly_doc_name.setText(doctor['name'])
            self.poly_doc_speciality.setText(doctor['speciality'])
            self.poly_doc_degree.setText(doctor.get('degree', ''))
            self.poly_doc_fees.setValue(doctor.get('visiting_fees', 0))
            
            # Load availability - organize by day
            availability = polyclinic_db.get_doctor_availability(doctor_id)
            
            # Group availability by day
            day_schedule = {}
            for avail in availability:
                day_idx = avail['day_of_week']
                if day_idx not in day_schedule:
                    day_schedule[day_idx] = []
                day_schedule[day_idx].append({
                    'start': avail['start_time'],
                    'end': avail['end_time']
                })
            
            # Clear and rebuild day widgets
            for day_idx in range(7):
                day_widget = self.poly_day_widgets[day_idx]
                
                # Clear existing slots
                day_widget['slots'].clear()
                while day_widget['slots_layout'].count() > 0:
                    item = day_widget['slots_layout'].takeAt(0)
                    if item:
                        item.deleteLater()
                
                # If this day has schedule, check it and add slots
                if day_idx in day_schedule:
                    day_widget['checkbox'].setChecked(True)
                    day_widget['add_btn'].setEnabled(True)  # Explicitly enable button
                    for slot in day_schedule[day_idx]:
                        self.poly_add_day_time_slot(day_idx, slot['start'], slot['end'])
                else:
                    day_widget['checkbox'].setChecked(False)
                    day_widget['add_btn'].setEnabled(False)  # Explicitly disable button
            
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Error", f"Error loading doctor: {str(e)}")
    
    def poly_delete_doctor(self, doctor_id):
        """Delete a doctor"""
        reply = QtWidgets.QMessageBox.question(self, "Confirm Delete", "Are you sure you want to delete this doctor?")
        if reply == QtWidgets.QMessageBox.StandardButton.Yes:
            try:
                polyclinic_db.delete_doctor(doctor_id)
                QtWidgets.QMessageBox.information(self, "Success", "Doctor deleted successfully!")
                self.poly_reload_doctor_list()
            except Exception as e:
                QtWidgets.QMessageBox.warning(self, "Error", f"Error deleting doctor: {str(e)}")
    
    def init_poly_queue_tab(self):
        """Initialize Patient Queue View tab"""
        main_layout = QtWidgets.QHBoxLayout(self.poly_queue_tab)
        
        # ===== LEFT PANEL: FILTERS =====
        left_widget = QtWidgets.QWidget()
        left_layout = QtWidgets.QVBoxLayout(left_widget)
        left_layout.setSpacing(10)
        
        # Filter group
        filter_group = QtWidgets.QGroupBox("Filters")
        filter_group_layout = QtWidgets.QFormLayout(filter_group)
        
        self.poly_queue_doctor_filter = QtWidgets.QComboBox()
        self.poly_queue_doctor_filter.currentIndexChanged.connect(self.poly_reload_queue)
        filter_group_layout.addRow("Doctor:", self.poly_queue_doctor_filter)
        
        self.poly_queue_date_filter = QtWidgets.QDateEdit()
        self.poly_queue_date_filter.setDate(QtCore.QDate.currentDate())
        self.poly_queue_date_filter.dateChanged.connect(self.poly_reload_queue)
        filter_group_layout.addRow("Date:", self.poly_queue_date_filter)
        
        self.poly_queue_time_filter = QtWidgets.QComboBox()
        self.poly_queue_time_filter.currentIndexChanged.connect(self.poly_reload_queue)
        filter_group_layout.addRow("Time Slot:", self.poly_queue_time_filter)
        
        left_layout.addWidget(filter_group)
        
        # Refresh button
        refresh_btn = QtWidgets.QPushButton("🔄 Refresh")
        MainWindow.style_button_with_dynamic_spacing(refresh_btn, font_size=11, padding="8px 16px")
        refresh_btn.clicked.connect(self.poly_reload_queue)
        left_layout.addWidget(refresh_btn)
        
        # Summary section
        summary_group = QtWidgets.QGroupBox("Summary")
        summary_group_layout = QtWidgets.QFormLayout(summary_group)
        
        self.poly_summary_total = QtWidgets.QLineEdit()
        self.poly_summary_total.setReadOnly(True)
        summary_group_layout.addRow("Total Patients:", self.poly_summary_total)
        
        self.poly_summary_attended = QtWidgets.QLineEdit()
        self.poly_summary_attended.setReadOnly(True)
        summary_group_layout.addRow("Attended:", self.poly_summary_attended)
        
        self.poly_summary_paid = QtWidgets.QLineEdit()
        self.poly_summary_paid.setReadOnly(True)
        summary_group_layout.addRow("Paid:", self.poly_summary_paid)
        
        left_layout.addWidget(summary_group)
        
        # Export Day button
        self.poly_export_day_btn = QtWidgets.QPushButton('📊 Export Day Data')
        self.poly_export_day_btn.setStyleSheet('font-size: 11px; padding: 8px 16px; background-color: #4CAF50; color: white; font-weight: bold;')
        self.poly_export_day_btn.clicked.connect(self.poly_export_day)
        left_layout.addWidget(self.poly_export_day_btn)
        
        left_layout.addStretch()
        
        # ===== RIGHT PANEL: TABLE =====
        right_widget = QtWidgets.QWidget()
        right_layout = QtWidgets.QVBoxLayout(right_widget)
        
        # Patient queue table
        self.poly_queue_table = QtWidgets.QTableWidget()
        self.poly_queue_table.setColumnCount(8)
        self.poly_queue_table.setHorizontalHeaderLabels(QUEUE_TABLE_HEADERS[:-1])  # Exclude 'Delete' column
        right_layout.addWidget(self.poly_queue_table)
        
        # Add panels to main layout
        main_layout.addWidget(left_widget, 0)  # Fixed width on left
        main_layout.addWidget(right_widget, 1)  # Expandable table on right
        
        # Don't load on init - lazy load on first focus
        self.poly_queue_loaded = False
    
    def poly_reload_queue(self):
        """Reload patient queue with filtering"""
        try:
            doctors = polyclinic_db.get_all_doctors()
            
            # Update doctor filter
            self.poly_queue_doctor_filter.blockSignals(True)
            current_doctor = self.poly_queue_doctor_filter.currentData()
            self.poly_queue_doctor_filter.clear()
            self.poly_queue_doctor_filter.addItem("All Doctors", None)
            for doc in doctors:
                self.poly_queue_doctor_filter.addItem(doc['name'], doc['doctor_id'])
            # Restore previous selection if it exists
            if current_doctor:
                idx = self.poly_queue_doctor_filter.findData(current_doctor)
                if idx >= 0:
                    self.poly_queue_doctor_filter.setCurrentIndex(idx)
            self.poly_queue_doctor_filter.blockSignals(False)
            
            # Get queue data
            selected_doctor_id = self.poly_queue_doctor_filter.currentData()
            selected_date = str(self.poly_queue_date_filter.date().toPython())
            selected_time = self.poly_queue_time_filter.currentText()
            
            if selected_doctor_id:
                bookings = polyclinic_db.get_bookings_for_doctor_date(selected_doctor_id, selected_date)
            else:
                bookings = []
                for doc in doctors:
                    bookings.extend(polyclinic_db.get_bookings_for_doctor_date(doc['doctor_id'], selected_date))
            
            # Collect unique time slots for filtering from doctor availability
            time_slots = set()
            selected_doctor_id = self.poly_queue_doctor_filter.currentData()
            
            if selected_doctor_id:
                # Get availability for selected doctor
                availability = polyclinic_db.get_doctor_availability(selected_doctor_id)
                for avail in availability:
                    time_slots.add(f"{avail['start_time']} - {avail['end_time']}")
            else:
                # Get availability for all doctors
                doctors = polyclinic_db.get_all_doctors()
                for doctor in doctors:
                    availability = polyclinic_db.get_doctor_availability(doctor['doctor_id'])
                    for avail in availability:
                        time_slots.add(f"{avail['start_time']} - {avail['end_time']}")
            
            # Update time filter
            self.poly_queue_time_filter.blockSignals(True)
            current_time = self.poly_queue_time_filter.currentText()
            self.poly_queue_time_filter.clear()
            self.poly_queue_time_filter.addItem("All Times")
            for ts in sorted(time_slots):
                self.poly_queue_time_filter.addItem(ts)
            # Restore previous selection if it exists
            if current_time:
                idx = self.poly_queue_time_filter.findText(current_time)
                if idx >= 0:
                    self.poly_queue_time_filter.setCurrentIndex(idx)
            self.poly_queue_time_filter.blockSignals(False)
            
            # Apply time filter if selected
            if selected_time and selected_time != "All Times":
                # Extract start time from selected time slot (format: "HH:MM - HH:MM")
                filtered_bookings = []
                for b in bookings:
                    booking_time = b.get('booking_time', '')
                    # Check if booking time falls within the selected time slot
                    if selected_time in booking_time or booking_time in selected_time:
                        filtered_bookings.append(b)
                bookings = filtered_bookings
            
            self.poly_queue_table.setRowCount(len(bookings))
            for r, booking in enumerate(bookings):
                patient = patient_cms_db.get_patient(booking['patient_id'])
                doctor = polyclinic_db.get_doctor(booking['doctor_id'])
                
                self.poly_queue_table.setItem(r, 0, QtWidgets.QTableWidgetItem(str(booking['serial_number'])))
                self.poly_queue_table.setItem(r, 1, QtWidgets.QTableWidgetItem(patient.get('name', '') if patient else ''))
                self.poly_queue_table.setItem(r, 2, QtWidgets.QTableWidgetItem(str(booking['patient_id'])))
                self.poly_queue_table.setItem(r, 3, QtWidgets.QTableWidgetItem(doctor.get('name', '') if doctor else ''))
                self.poly_queue_table.setItem(r, 4, QtWidgets.QTableWidgetItem(booking.get('booking_time', '')))
                self.poly_queue_table.setItem(r, 5, QtWidgets.QTableWidgetItem(patient.get('phone', '') if patient else ''))
                
                # Payment checkbox - block signals, set state, connect signal
                payment_container = QtWidgets.QWidget()
                payment_layout = QtWidgets.QHBoxLayout(payment_container)
                payment_layout.setContentsMargins(0, 0, 0, 0)
                payment_check = QtWidgets.QCheckBox()
                payment_check.blockSignals(True)
                payment_check.setChecked(booking['payment_status'] == 'PAID')
                payment_check.blockSignals(False)
                payment_check.stateChanged.connect(lambda state, bid=booking['booking_id']: self.poly_update_payment(bid, state == 2))
                payment_layout.addStretch()
                payment_layout.addWidget(payment_check)
                payment_layout.addStretch()
                self.poly_queue_table.setCellWidget(r, 6, payment_container)
                
                # Attendance checkbox - block signals, set state, connect signal
                attendance_container = QtWidgets.QWidget()
                attendance_layout = QtWidgets.QHBoxLayout(attendance_container)
                attendance_layout.setContentsMargins(0, 0, 0, 0)
                attendance_check = QtWidgets.QCheckBox()
                attendance_check.blockSignals(True)
                attendance_check.setChecked(booking['attendance_status'] == 'ATTENDED')
                attendance_check.blockSignals(False)
                attendance_check.stateChanged.connect(lambda state, bid=booking['booking_id']: self.poly_update_attendance(bid, state == 2))
                attendance_layout.addStretch()
                attendance_layout.addWidget(attendance_check)
                attendance_layout.addStretch()
                self.poly_queue_table.setCellWidget(r, 7, attendance_container)
            
            # Update summary
            if selected_doctor_id:
                summary = polyclinic_db.get_day_summary(selected_doctor_id, selected_date)
            else:
                summary = {'total_patients': len(bookings), 'attended': 0, 'paid': 0}
                for booking in bookings:
                    if booking['attendance_status'] == 'ATTENDED':
                        summary['attended'] += 1
                    if booking['payment_status'] == 'PAID':
                        summary['paid'] += 1
            
            self.poly_summary_total.setText(str(summary.get('total_patients', 0)))
            self.poly_summary_attended.setText(str(summary.get('attended', 0)))
            self.poly_summary_paid.setText(str(summary.get('paid', 0)))
            
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Error", f"Error loading queue: {str(e)}")
    
    def poly_update_payment(self, booking_id, is_checked):
        """Update payment status"""
        try:
            status = 'PAID' if is_checked else 'PENDING'
            polyclinic_db.update_booking_payment_status(booking_id, status)
            self.poly_reload_queue()
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Error", f"Error updating payment: {str(e)}")
    
    def poly_update_attendance(self, booking_id, is_checked):
        """Update attendance status"""
        try:
            status = 'ATTENDED' if is_checked else 'PENDING'
            polyclinic_db.update_booking_attendance_status(booking_id, status)
            self.poly_reload_queue()
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Error", f"Error updating attendance: {str(e)}")
    
    def poly_delete_booking(self, booking_id):
        """Delete a booking"""
        reply = QtWidgets.QMessageBox.question(self, "Confirm Delete", "Are you sure?")
        if reply == QtWidgets.QMessageBox.StandardButton.Yes:
            try:
                polyclinic_db.delete_booking(booking_id)
                self.poly_reload_queue()
            except Exception as e:
                QtWidgets.QMessageBox.warning(self, "Error", f"Error deleting booking: {str(e)}")
    
    def poly_export_day(self):
        """Export day data to XLSX"""
        try:
            selected_doctor_id = self.poly_queue_doctor_filter.currentData()
            selected_date = str(self.poly_queue_date_filter.date().toPython())
            
            if not selected_doctor_id:
                QtWidgets.QMessageBox.warning(self, "Error", "Please select a doctor for export")
                return
            
            doctor = polyclinic_db.get_doctor(selected_doctor_id)
            
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Queue"
            
            # Headers
            headers = ['Serial', 'Patient Name', PATIENT_ID_LABEL, 'Phone', 'Payment', 'Attendance', 'Fees']
            ws.append(headers)
            
            # Style header
            header_fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Get bookings
            bookings = polyclinic_db.get_bookings_for_doctor_date(selected_doctor_id, selected_date)
            
            total_fees = 0
            collected_fees = 0
            for booking in bookings:
                patient = patient_cms_db.get_patient(booking['patient_id'])
                fees = doctor.get('visiting_fees', 0)
                total_fees += fees
                if booking['payment_status'] == 'PAID':
                    collected_fees += fees
                
                ws.append([
                    booking['serial_number'],
                    patient.get('name', '') if patient else '',
                    booking['patient_id'],
                    patient.get('phone', '') if patient else '',
                    booking['payment_status'],
                    booking['attendance_status'],
                    fees
                ])
            
            # Summary rows
            ws.append([])
            ws.append(['SUMMARY'])
            ws.append(['Total Patients:', len(bookings)])
            ws.append(['Total Fees:', total_fees])
            ws.append(['Collected Fees:', collected_fees])
            ws.append(['Pending Fees:', total_fees - collected_fees])
            
            # Save file
            filename = f"Polyclinic_Queue_{doctor['name']}_{selected_date}.xlsx"
            wb.save(filename)
            
            QtWidgets.QMessageBox.information(self, "Success", f"Data exported to {filename}")
            
        except ImportError:
            QtWidgets.QMessageBox.warning(self, "Error", "openpyxl not installed. Please install it to use export feature.")
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Error", f"Error exporting: {str(e)}")
    
    def init_poly_cms_tab(self):
        """Initialize Patient CMS tab for Polyclinic (shared with Pathology)"""
        # Reuse the same patient CMS implementation as pathology
        self.init_cms_tab_impl(self.poly_cms_tab)


def main():
    patient_cms_db.init_db()
    datasheet_db.init_db()
    report_tracker_db.init_db()
    auth_db.init_db()
    catalogue_db.init_db()
    special_tests_db.init_db()
    
    app = QtWidgets.QApplication(sys.argv)
    app.setQuitOnLastWindowClosed(False)
    
    # Apply modern theme
    theme.apply_stylesheet(app)
    
    try:
        data_fetcher.start_fetch_scheduler()
    except Exception:
        pass
    
    if not auth_db.check_if_any_user_exists():
        # Show branded First Time Setup
        setup_dlg = FirstTimeSetupDialog()
        if setup_dlg.exec() != QtWidgets.QDialog.Accepted:
            # User cancelled setup
            sys.exit(0)
    
    login_win = LoginWindow()
    main_window = None
    
    def on_login(user):
        nonlocal main_window
        main_window = MainWindow(user)
        main_window.logout_signal.connect(lambda: show_login(user))
        main_window.show()
        login_win.hide()
    
    def show_login(user):
        nonlocal main_window
        if main_window:
            main_window.close()
            main_window = None
        login_win.show()
        login_win.username.clear()
        login_win.password.clear()
    
    login_win.logged_in.connect(on_login)
    login_win.show()
    
    sys.exit(app.exec())


def global_exception_handler(exctype, value, traceback_obj):
    """
    Global exception handler to ensure the app doesn't crash silently.
    Shows a critical error message box with the exception details.
    """
    import traceback
    
    # Ignore keyboard interrupt (Ctrl+C)
    if exctype == KeyboardInterrupt:
        sys.__excepthook__(exctype, value, traceback_obj)
        return
        
    error_msg = "".join(traceback.format_exception(exctype, value, traceback_obj))
    print(f"CRITICAL ERROR: {error_msg}", file=sys.stderr)
    
    # Ensure we utilize the main thread's QApplication if possible
    app = QtWidgets.QApplication.instance()
    if app:
        error_box = QtWidgets.QMessageBox()
        error_box.setIcon(QtWidgets.QMessageBox.Critical)
        error_box.setWindowTitle("Critical Application Error")
        error_box.setText("An unexpected error occurred. The application must close.")
        error_box.setDetailedText(error_msg)
        error_box.setStandardButtons(QtWidgets.QMessageBox.Ok)
        error_box.exec()
        
    sys.exit(1)


if __name__ == '__main__':
    # usage of hook
    sys.excepthook = global_exception_handler
    main()
